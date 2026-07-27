"""
Genera mapa_muestras.html + carpeta imagenes/ con toda la informacion de muestras geologicas.
Incluye pagina de Geoquimica con diagramas por tipo de roca.

Uso:
    python generar_mapa.py "ruta\\archivo.xlsx" [salida.html]
                           [--fotos "ruta\\fotos"]
                           [--new-photos "ruta\\new_photos"]
                           [--laminas "ruta\\laminas delgadas"]

Requiere:  pip install openpyxl matplotlib
Opcional:  pip install Pillow
"""

import sys
import json
import re
import shutil
import math
import warnings
import csv
import subprocess
from pathlib import Path

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instala con: pip install openpyxl")

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from matplotlib.patches import Polygon as MplPolygon
    from matplotlib.collections import PatchCollection
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

try:
    import pandas as pd
    from pyrolite.mineral.normative import CIPW_norm
    HAS_PYROLITE = True
except ImportError:
    HAS_PYROLITE = False

try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# ---------------------------------------------------------------------------
# Configuracion
# ---------------------------------------------------------------------------
SHEET_MAIN  = "Hoja1"
SHEET_DRX   = "DRX_EVA"
SHEET_GEO   = "Geoquim"
MAX_IMG_PX  = 1400
IMG_QUALITY = 85

COLUMN_KEYS = [
    "UNINORTE_CODE", "DESTRUCTIVE", "GARDEN_CODE", "PAIS", "UBICACION",
    "SPP", "FECHA", "LATITUD", "LONGITUD", "AUTOR", "TIPO_ROCA",
    "CLASIFICACION", "PESO_GEOQ", "OBS_PETRO", "GEOQUIMICA", "DRX",
    "PESO_TOTAL",
]

# Pesos moleculares
MW = {
    "SiO2": 60.09, "TiO2": 79.90, "Al2O3": 101.96, "Fe2O3": 159.69,
    "FeO": 71.85, "MnO": 70.94, "MgO": 40.30, "CaO": 56.08,
    "Na2O": 61.98, "K2O": 94.20, "P2O5": 141.94
}

SCRIPT_DIR = Path(__file__).parent

# ---------------------------------------------------------------------------
# Lectura del Excel — Hoja1
# ---------------------------------------------------------------------------
def leer_hoja_principal(ws):
    filas = list(ws.iter_rows(values_only=True))
    datos = []
    for fila in filas[1:]:
        if all(v is None for v in fila):
            continue
        fila = list(fila[: len(COLUMN_KEYS)])
        while len(fila) < len(COLUMN_KEYS):
            fila.append(None)
        fila = [v.isoformat() if hasattr(v, "isoformat") else v for v in fila]
        datos.append(fila)
    return datos


# ---------------------------------------------------------------------------
# Lectura del Excel — DRX_EVA
# ---------------------------------------------------------------------------
def parsear_drx_eva(ws):
    result = {}
    rows = list(ws.iter_rows(values_only=True))
    i = 0
    while i < len(rows):
        row = rows[i]
        raw_code = row[0]
        if raw_code and re.match(r'NYB\d+', str(raw_code).strip(), re.IGNORECASE):
            code = str(raw_code).strip().upper()
            m = re.match(r'NYB(\d+)', code, re.IGNORECASE)
            if m:
                code = f"NYB{int(m.group(1)):02d}"
            block = {"cod_ids": [], "minerales": [], "formulas": [],
                     "descripcion": "", "interpretacion": "", "nota": ""}
            has_nota_col = len(row) > 8 and row[8] == 'NOTA DRX EVA'
            block["cod_ids"] = [
                str(v).strip() for v in row[2:]
                if v is not None and str(v).strip() and 'NOTA' not in str(v)
            ]
            i += 1
            while i < len(rows):
                r = rows[i]
                if r[0] is not None and re.match(r'NYB\d+', str(r[0]).strip(), re.IGNORECASE):
                    break
                label = str(r[1]).strip().lower() if r[1] else ""
                if "mineral" in label:
                    block["minerales"] = [str(v).strip() for v in r[2:8] if v is not None and str(v).strip()]
                    if has_nota_col and len(r) > 8 and r[8]:
                        block["nota"] = str(r[8]).strip()
                elif "formula" in label or "quimica" in label:
                    block["formulas"] = [str(v).strip() for v in r[2:8] if v is not None and str(v).strip()]
                elif "descripci" in label:
                    if r[2]: block["descripcion"] = str(r[2]).strip()
                elif "interpretaci" in label:
                    if r[2]: block["interpretacion"] = str(r[2]).strip()
                i += 1
            result[code] = block
        else:
            i += 1
    return result


# ---------------------------------------------------------------------------
# Lectura del Excel — Geoquim
# ---------------------------------------------------------------------------
def leer_geoquim(ws):
    """
    Devuelve lista de dicts con la geoquimica de cada muestra.
    Maneja valores '< 1' (bajo limite de deteccion) -> 0.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Fila 1 (indice 0): nombres de analitos
    # Fila 2 (indice 1): unidades
    # Datos desde fila 5 (indice 4) aprox — buscar primer NYB
    headers = [str(h).strip() if h else "" for h in rows[0]]

    data = []
    for row in rows[1:]:
        if not row[0]:
            continue
        code_raw = str(row[0]).strip()
        m = re.match(r'NYB(\d+)', code_raw, re.IGNORECASE)
        if not m:
            continue
        code = f"NYB{int(m.group(1)):02d}"
        rec = {"UNINORTE_CODE": code}
        for j, h in enumerate(headers[1:], 1):
            if not h or j >= len(row):
                continue
            v = row[j]
            if v is None:
                rec[h] = None
            elif isinstance(v, str) and '<' in v:
                rec[h] = 0.0
            else:
                try:
                    rec[h] = float(v)
                except (ValueError, TypeError):
                    rec[h] = None
        data.append(rec)
    return data


# ---------------------------------------------------------------------------
# Recopilacion de laminas delgadas
# ---------------------------------------------------------------------------
IMG_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.JPG', '.JPEG', '.PNG'}

def es_imagen(path: Path) -> bool:
    return path.suffix in IMG_EXTENSIONS

def resize_copy(src: Path, dst: Path):
    if HAS_PIL:
        try:
            with PILImage.open(src) as img:
                img = img.convert("RGB")
                img.thumbnail((MAX_IMG_PX, MAX_IMG_PX), PILImage.LANCZOS)
                img.save(dst, "JPEG", quality=IMG_QUALITY, optimize=True)
            return
        except Exception:
            pass
    shutil.copy2(src, dst)


def recopilar_laminas_delgadas(all_codes, laminas_dir, img_out_dir: Path):
    """
    Busca pares NC/NP en la carpeta de laminas delgadas.
    Patron: {CODE}_{N}NC.ext y {CODE}_{N}NP.ext
    Devuelve dict: {code: [{"nc": ruta, "np": ruta}, ...]}
    """
    result = {c: [] for c in all_codes}
    if not laminas_dir:
        return result
    laminas_path = Path(laminas_dir)
    if not laminas_path.is_dir():
        print(f"  Aviso: carpeta laminas no encontrada: {laminas_dir}")
        return result

    lam_out = img_out_dir / "laminas"
    lam_out.mkdir(parents=True, exist_ok=True)

    # Recopilar todos los archivos recursivamente
    pat_nc = re.compile(r'(NYB\d+)_(\d+)NC', re.IGNORECASE)
    pat_np = re.compile(r'(NYB\d+)_(\d+)NP', re.IGNORECASE)

    nc_files = {}  # (code, num) -> Path
    np_files = {}

    for f in sorted(laminas_path.rglob("*")):
        if not f.is_file() or not es_imagen(f):
            continue
        mn = pat_nc.search(f.stem)
        mp = pat_np.search(f.stem)
        if mn:
            code = f"NYB{int(mn.group(1)[3:]):02d}"
            num  = int(mn.group(2))
            nc_files[(code, num)] = f
        elif mp:
            code = f"NYB{int(mp.group(1)[3:]):02d}"
            num  = int(mp.group(2))
            np_files[(code, num)] = f

    # Emparejar NC/NP
    keys = set(nc_files.keys()) | set(np_files.keys())
    from collections import defaultdict
    by_code = defaultdict(list)
    for (code, num) in keys:
        by_code[code].append(num)

    for code, nums in by_code.items():
        if code not in result:
            continue
        for num in sorted(set(nums)):
            nc_src = nc_files.get((code, num))
            np_src = np_files.get((code, num))
            nc_dst_rel = np_dst_rel = None

            if nc_src:
                dst_name = f"{code}_{num}NC.jpg"
                dst = lam_out / dst_name
                if not dst.exists():
                    resize_copy(nc_src, dst)
                nc_dst_rel = f"imagenes/laminas/{dst_name}"

            if np_src:
                dst_name = f"{code}_{num}NP.jpg"
                dst = lam_out / dst_name
                if not dst.exists():
                    resize_copy(np_src, dst)
                np_dst_rel = f"imagenes/laminas/{dst_name}"

            result[code].append({"nc": nc_dst_rel, "np": np_dst_rel, "num": num})

    total = sum(1 for v in result.values() if v)
    print(f"  -> {total} muestras con laminas delgadas")
    return result


# ---------------------------------------------------------------------------
# Imagenes de muestra de mano
# ---------------------------------------------------------------------------
def recopilar_imagenes_muestra_v2(all_codes, fotos_dir, new_photos_dir, img_out_dir):
    result = {c: [] for c in all_codes}
    img_out_dir.mkdir(parents=True, exist_ok=True)

    # Primero: recoger imagenes ya procesadas en img_out_dir (NYBxx_muestra_N.jpg)
    import re as _re
    for code in all_codes:
        existing = sorted(
            f for f in img_out_dir.glob(f"{code}_muestra_*.jpg")
        )
        if existing:
            result[code] = [f"imagenes/{f.name}" for f in existing]

    if fotos_dir:
        fotos_base = Path(fotos_dir)
        if fotos_base.is_dir():
            for code in all_codes:
                sub = fotos_base / code
                if not sub.is_dir():
                    continue
                imgs = sorted([f for f in sub.iterdir() if es_imagen(f)])
                for idx, src in enumerate(imgs):
                    dst_name = f"{code}_muestra_{idx}.jpg"
                    dst = img_out_dir / dst_name
                    if not dst.exists():
                        resize_copy(src, dst)
                    result[code].append(f"imagenes/{dst_name}")
        else:
            print(f"  Aviso: carpeta fotos no encontrada: {fotos_dir}")

    if new_photos_dir:
        new_base = Path(new_photos_dir)
        if new_base.is_dir():
            pattern = re.compile(r'NYB(\d+)', re.IGNORECASE)
            buckets = {c: [] for c in all_codes}
            for f in sorted(new_base.iterdir()):
                if not es_imagen(f):
                    continue
                nums = [int(x) for x in pattern.findall(f.stem)]
                for n in nums:
                    code = f"NYB{n:02d}"
                    if code in buckets:
                        buckets[code].append(f)
            for code, srcs in buckets.items():
                existing = len(result[code])
                for src in sorted(srcs):
                    dst_name = f"{code}_muestra_{existing}.jpg"
                    dst = img_out_dir / dst_name
                    if not dst.exists():
                        resize_copy(src, dst)
                    result[code].append(f"imagenes/{dst_name}")
                    existing += 1
        else:
            print(f"  Aviso: carpeta new_photos no encontrada: {new_photos_dir}")

    return result


def recopilar_imagenes_drx(all_codes, img_out_dir: Path):
    drx_dir = SCRIPT_DIR / "imagenes_drx"
    result = {c: None for c in all_codes}
    img_out_dir.mkdir(parents=True, exist_ok=True)
    if not drx_dir.is_dir():
        return result
    pattern = re.compile(r'NYB(\d+)', re.IGNORECASE)
    for f in sorted(drx_dir.iterdir()):
        if not es_imagen(f):
            continue
        m = pattern.search(f.stem)
        if m:
            code = f"NYB{int(m.group(1)):02d}"
            if code in result:
                dst_name = f"{code}_drx.jpg"
                dst = img_out_dir / dst_name
                if not dst.exists():
                    resize_copy(f, dst)
                result[code] = f"imagenes/{dst_name}"
    return result


# ---------------------------------------------------------------------------
# DIAGRAMAS GEOQUIMICOS
# ---------------------------------------------------------------------------
def _safe(v, default=0.0):
    if v is None: return default
    try:
        f = float(v)
        return f if math.isfinite(f) else default
    except (ValueError, TypeError):
        return default

def _mol(v, oxide):
    return _safe(v) / MW.get(oxide, 1.0)


def _ternary_coords(a, b, c):
    """Convierte coordenadas ternarias (a,b,c deben sumar 1 o se normalizan) a (x,y)."""
    s = a + b + c
    if s == 0:
        return None, None
    a, b, c = a/s, b/s, c/s
    x = b + c * 0.5
    y = c * math.sqrt(3) / 2
    return x, y


def _draw_ternary_frame(ax, labels, fontsize=9):
    """Dibuja el triangulo ternario y etiquetas en ejes normales."""
    tri = plt.Polygon([[0,0],[1,0],[0.5, math.sqrt(3)/2]], fill=False,
                       edgecolor='black', linewidth=1.5)
    ax.add_patch(tri)
    ax.set_xlim(-0.1, 1.1)
    ax.set_ylim(-0.1, 1.0)
    ax.set_aspect('equal')
    ax.axis('off')
    # Etiquetas de vertices
    ax.text(0,   -0.07, labels[0], ha='center', va='top',  fontsize=fontsize, fontweight='bold')
    ax.text(1,   -0.07, labels[1], ha='center', va='top',  fontsize=fontsize, fontweight='bold')
    ax.text(0.5, math.sqrt(3)/2+0.04, labels[2], ha='center', va='bottom', fontsize=fontsize, fontweight='bold')


def _cipw_dataframe(data):
    """Construye DataFrame para CIPW_norm de pyrolite."""
    rows = []
    for r in data:
        row = {
            "SiO2":  _safe(r.get("SiO2")),
            "TiO2":  _safe(r.get("TiO2")),
            "Al2O3": _safe(r.get("Al2O3")),
            "FeOT":  _safe(r.get("Fe2O3(T)")) * 0.8998,  # total Fe como FeO
            "MnO":   _safe(r.get("MnO")),
            "MgO":   _safe(r.get("MgO")),
            "CaO":   _safe(r.get("CaO")),
            "Na2O":  _safe(r.get("Na2O")),
            "K2O":   _safe(r.get("K2O")),
            "P2O5":  _safe(r.get("P2O5")),
        }
        rows.append(row)
    df = pd.DataFrame(rows, index=[r["UNINORTE_CODE"] for r in data])
    return df.astype(float)


def generar_diagramas(filas, geochem_data, tipo_roca_map, out_dir: Path):
    """
    Genera diagramas PNG y devuelve dict de rutas relativas.
    tipo_roca_map: {code: tipo_roca}
    """
    if not HAS_MPL:
        print("  matplotlib no instalado, omitiendo diagramas.")
        return {}

    out_dir.mkdir(parents=True, exist_ok=True)
    rutas = {}

    # Separar por tipo de roca
    igneas = [r for r in geochem_data if tipo_roca_map.get(r["UNINORTE_CODE"], "").upper() == "IGNEA"]
    seds   = [r for r in geochem_data if tipo_roca_map.get(r["UNINORTE_CODE"], "").upper() == "SEDIMENTARIA"]
    metas  = [r for r in geochem_data if tipo_roca_map.get(r["UNINORTE_CODE"], "").upper() == "METAMORFICA"]

    print(f"  Geoquimica: {len(igneas)} igneas, {len(seds)} sedimentarias, {len(metas)} metamorficas")

    # Calcular norma CIPW para igneas si pyrolite disponible
    cipw_norm = None
    if igneas and HAS_PYROLITE:
        try:
            import warnings
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                df_cipw = _cipw_dataframe(igneas)
                cipw_norm = CIPW_norm(df_cipw)
            print("  CIPW normativo calculado con pyrolite.")
        except Exception as e:
            print(f"  Aviso CIPW: {e}")

    # ---------- IGNEAS ----------
    if igneas:
        rutas["igneas_tas"]      = _diagrama_tas(igneas, out_dir)
        rutas["igneas_qapf"]     = _diagrama_qapf(igneas, cipw_norm, out_dir)
        rutas["igneas_afm"]      = _diagrama_afm(igneas, out_dir)
        rutas["igneas_harker"]   = _diagrama_harker(igneas, out_dir)
        rutas["igneas_shervais"] = _diagrama_shervais(igneas, out_dir)
        rutas = {k: v for k, v in rutas.items() if v}

    # ---------- SEDIMENTARIAS ----------
    if seds:
        rutas["sed_cia"]    = _diagrama_cia(seds, out_dir)
        rutas["sed_herron"] = _diagrama_herron(seds, out_dir)
        rutas["sed_acnk"]   = _diagrama_acnk(seds, out_dir)

    # ---------- METAMORFICAS ----------
    if metas:
        rutas["meta_wf"]  = _diagrama_winchester_floyd(metas, out_dir)
        rutas["meta_zrti"]= _diagrama_zr_ti(metas, out_dir)
        rutas["meta_acf"] = _diagrama_acf(metas, out_dir)

    plt.close('all')
    return {k: v for k, v in rutas.items() if v}


# ---- TAS (Total Alkali-Silica) ----
def _diagrama_tas(data, out_dir):
    fig, ax = plt.subplots(figsize=(8, 6))

    # Boundaries simplificadas de Le Bas et al. 1986
    # Lineas principales del diagrama TAS
    tas_lines = [
        # Foidita -> Tefrita/Basanita
        ([41,45],[0,5],'gray'),
        ([41,41],[0,3],'gray'),
        # Basalto -> Andesita
        ([45,52],[5,5],'gray'),
        ([52,57],[5,5],'gray'),
        # Limites verticales
        ([45,45],[0,5],'gray'),
        ([52,52],[0,5],'gray'),
        ([57,57],[0,5.9],'gray'),
        ([63,63],[0,7],'gray'),
        ([69,69],[0,8],'gray'),
        # Division alcalino/subalcalino (Irvine & Baragar 1971)
        ([39,40,43,45,48.4,53,60,70],
         [0,0.4,2,2.8,4,6,8,8],'black'),
    ]

    # Poligonos de campo con colores suaves
    campos = {
        'Foidita':       {'xy':[[35,0],[41,0],[41,7],[35,9]],'c':'#d0e8ff'},
        'Picro-\nbasalto':{'xy':[[41,0],[45,0],[45,3],[41,3]],'c':'#c5e8c5'},
        'Basalto':       {'xy':[[45,0],[52,0],[52,5],[45,5]],'c':'#b0d8b0'},
        'Andesita\nbasaltica':{'xy':[[52,0],[57,0],[57,5.9],[52,5]],'c':'#ffe0a0'},
        'Andesita':      {'xy':[[57,0],[63,0],[63,7],[57,5.9]],'c':'#ffd070'},
        'Dacita':        {'xy':[[63,0],[69,0],[69,8],[63,7]],'c':'#ffb060'},
        'Riolita':       {'xy':[[69,0],[78,0],[78,8],[69,8]],'c':'#ff9060'},
        'Traquibasalto': {'xy':[[45,5],[52,5],[53,9],[49,9.3],[45,9.4]],'c':'#c0f0c0'},
        'Traquiandesita\nbasaltica':{'xy':[[52,5],[57,5.9],[57,11.5],[53,9]],'c':'#a8e8a8'},
        'Traquiandesita':{'xy':[[57,5.9],[63,7],[63,11.5],[57,11.5]],'c':'#90d090'},
        'Traquita':      {'xy':[[63,7],[69,8],[69,13],[63,11.5]],'c':'#78b878'},
        'Fonolita':      {'xy':[[53,9],[57,11.5],[63,11.5],[57,14],[53,14]],'c':'#b0c8ff'},
        'Tefrita/\nBasanita':{'xy':[[41,3],[45,3],[45,9.4],[49,9.3],[45,5],[41,7]],'c':'#d8d8ff'},
    }

    for nombre, info in campos.items():
        poly = MplPolygon(info['xy'], closed=True, fc=info['c'], ec='gray',
                          alpha=0.5, linewidth=0.7)
        ax.add_patch(poly)
        xs = [p[0] for p in info['xy']]
        ys = [p[1] for p in info['xy']]
        cx, cy = sum(xs)/len(xs), sum(ys)/len(ys)
        ax.text(cx, cy, nombre, ha='center', va='center',
                fontsize=6, color='#333', style='italic')

    # Linea alcalino/subalcalino
    line_x = [39,40,43,45,48.4,53,60,70,78]
    line_y = [0, 0.4,2, 2.8,4,  6, 8, 8, 8]
    ax.plot(line_x, line_y, 'k--', linewidth=0.8, alpha=0.6, label='Irvine & Baragar (1971)')

    # Puntos
    COLORS = {'IGNEA':'#e63946','SEDIMENTARIA':'#f4b400','METAMORFICA':'#2a9d3f'}
    for r in data:
        sio2 = _safe(r.get("SiO2"))
        na2o = _safe(r.get("Na2O"))
        k2o  = _safe(r.get("K2O"))
        tas  = na2o + k2o
        if sio2 > 0:
            ax.scatter(sio2, tas, s=60, color='#e63946', zorder=5,
                       edgecolors='white', linewidths=0.7)
            ax.annotate(r["UNINORTE_CODE"], (sio2, tas),
                        fontsize=6, ha='left', va='bottom',
                        xytext=(3, 2), textcoords='offset points')

    ax.set_xlim(35, 78)
    ax.set_ylim(0, 16)
    ax.set_xlabel("SiO₂ (%)", fontsize=10)
    ax.set_ylabel("Na₂O + K₂O (%)", fontsize=10)
    ax.set_title("Diagrama TAS — Clasificacion de Rocas Igneas\n(Le Bas et al. 1986)", fontsize=10)
    ax.grid(True, alpha=0.3, linestyle=':')
    ax.legend(fontsize=7, loc='upper left')

    fname = out_dir / "igneas_tas.png"
    fig.tight_layout()
    fig.savefig(fname, dpi=130, bbox_inches='tight')
    plt.close(fig)
    print(f"    -> {fname.name}")
    return f"imagenes/diagramas/igneas_tas.png"


# ---- QAPF normativo (pyrolite) ----
def _diagrama_qapf(data, cipw_norm, out_dir):
    """
    Diagrama QAPF usando minerales normativos CIPW (pyrolite).
    Si cipw_norm es None, omite el diagrama.
    """
    if cipw_norm is None:
        return None

    fig, ax = plt.subplots(figsize=(7, 6))

    # Triangulo QAPF superior (Q-A-P) — rocas sin feldespatoides
    # Vertices: Q(arriba)=0.5,sqrt(3)/2  A(abajo-izq)=0,0  P(abajo-der)=1,0
    h = math.sqrt(3) / 2
    tri = MplPolygon([[0,0],[1,0],[0.5,h]], closed=True,
                     fc='#f8f8f8', ec='black', linewidth=1.5)
    ax.add_patch(tri)

    # Lineas de campo QAPF (% de Q, A, P sobre el total Q+A+P)
    # Lineas horizontales de Q% = 5, 20, 60 (en coordenadas del triangulo)
    def qap_to_xy(q_pct, a_pct, p_pct):
        s = q_pct + a_pct + p_pct
        if s == 0: return None, None
        q, a, p = q_pct/s, a_pct/s, p_pct/s
        # Q en vertice top, A en bot-izq, P en bot-der
        x = a*0 + p*1 + q*0.5
        y = a*0 + p*0 + q*h
        return x, y

    # Lineas de campo
    for q_val in [5, 20, 60]:
        pts = [qap_to_xy(q_val, a, 100-q_val-a) for a in range(0, 101-q_val, 5)]
        pts = [(x,y) for x,y in pts if x is not None]
        if pts:
            ax.plot([p[0] for p in pts],[p[1] for p in pts],
                    color='gray', linewidth=0.7, alpha=0.6)

    for p_ratio in [10, 35, 65, 90]:  # % de P/(A+P)
        q_vals = [0, 5, 20, 60, 90]
        pts = []
        for q in q_vals:
            remain = 100 - q
            p = remain * p_ratio / 100
            a = remain - p
            x, y = qap_to_xy(q, a, p)
            if x: pts.append((x,y))
        if pts:
            ax.plot([p[0] for p in pts],[p[1] for p in pts],
                    color='gray', linewidth=0.7, alpha=0.6)

    # Etiquetas de campos
    campos_qapf = [
        ("Granito",             0.5,  0.5,  0.5,  0.5),   # (q%,a%,p%) centroide aprox
        ("Granodiorita",        0.25, 0.25, 0.75, 0.25),
        ("Tonalita",            0.25, 0.05, 0.95, 0.2),
        ("Sienita\ncuarzosa",   0.12, 0.7,  0.3,  0.08),
        ("Monzonita\ncuarzosa", 0.12, 0.45, 0.55, 0.08),
        ("Monzodiorita\ncuarzosa",0.1,0.2, 0.8,  0.07),
        ("Sienita",             0.03, 0.7,  0.3,  0.025),
        ("Monzonita",           0.03, 0.45, 0.55, 0.025),
        ("Diorita/Gabro",       0.03, 0.1,  0.9,  0.025),
    ]

    campo_colors = ['#ffe0d0','#ffd0b0','#ffc090','#e0d0ff','#d0c0ff',
                    '#c0b0ff','#d0ffe0','#c0ffcc','#b0ffb8']
    for (nombre, qa, aa, pa, _), col in zip(campos_qapf, campo_colors):
        x, y = qap_to_xy(qa*100, aa*100, pa*100)
        if x:
            ax.text(x, y, nombre, ha='center', va='center',
                    fontsize=6.5, color='#333', style='italic',
                    bbox=dict(fc=col, alpha=0.6, pad=2, ec='none', boxstyle='round'))

    # Etiquetas de vertices
    ax.text(0.5, h+0.04, "Q", ha='center', va='bottom', fontsize=12, fontweight='bold')
    ax.text(-0.06, -0.04, "A", ha='center', va='top',   fontsize=12, fontweight='bold')
    ax.text(1.06, -0.04,  "P", ha='center', va='top',   fontsize=12, fontweight='bold')

    # Porcentajes de escala
    for pct in [20, 40, 60, 80]:
        # Linea de Q%
        x1, y1 = qap_to_xy(pct, 100-pct, 0)
        x2, y2 = qap_to_xy(pct, 0, 100-pct)
        if x1 and x2:
            ax.plot([x1,x2],[y1,y2],'--',color='#aaa',linewidth=0.5,alpha=0.5)
            ax.text(x1-0.04, y1, f'{pct}', fontsize=6, color='#888', ha='right')

    # Plotear muestras con CIPW
    norm_cols = cipw_norm.columns.tolist()
    q_col  = next((c for c in norm_cols if 'quartz' in c.lower()), None)
    or_col = next((c for c in norm_cols if 'orthoclase' in c.lower()), None)
    ab_col = next((c for c in norm_cols if 'albite' in c.lower()), None)
    an_col = next((c for c in norm_cols if 'anorthite' in c.lower()), None)
    # Feldespatoides
    neph_col = next((c for c in norm_cols if 'nepheline' in c.lower()), None)
    leuc_col = next((c for c in norm_cols if 'leucite' in c.lower()), None)

    for r in data:
        code = r["UNINORTE_CODE"]
        if code not in cipw_norm.index:
            continue
        row = cipw_norm.loc[code]
        q  = _safe(row.get(q_col, 0)  if q_col  else 0)
        or_= _safe(row.get(or_col, 0) if or_col else 0)
        ab = _safe(row.get(ab_col, 0) if ab_col else 0)
        an = _safe(row.get(an_col, 0) if an_col else 0)
        f  = (_safe(row.get(neph_col, 0) if neph_col else 0) +
              _safe(row.get(leuc_col, 0)  if leuc_col  else 0))
        P = ab + an
        total = q + or_ + P
        if total < 0.01:
            continue
        x, y = qap_to_xy(q, or_, P)
        if x is not None:
            ax.scatter(x, y, s=70, color='#e63946', zorder=6,
                       edgecolors='white', linewidths=0.8)
            ax.annotate(code, (x, y), fontsize=6.5, ha='left', va='bottom',
                        xytext=(4, 3), textcoords='offset points',
                        bbox=dict(fc='white', alpha=0.6, pad=1, ec='none'))

    ax.set_xlim(-0.15, 1.15)
    ax.set_ylim(-0.1, h+0.12)
    ax.set_aspect('equal')
    ax.axis('off')
    ax.set_title("Diagrama QAPF Normativo (CIPW)\nStreckeisen (1976) — Rocas Igneas Plutónicas",
                 fontsize=10)

    fname = out_dir / "igneas_qapf.png"
    fig.tight_layout()
    fig.savefig(fname, dpi=140, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    print(f"    -> {fname.name}")
    return f"imagenes/diagramas/igneas_qapf.png"


# ---- AFM ternario ----
def _diagrama_afm(data, out_dir):
    fig, ax = plt.subplots(figsize=(6.5, 6))
    _draw_ternary_frame(ax, ['A\n(Na₂O+K₂O)', 'M\n(MgO)', 'F\n(FeO_total)'])
    ax.set_title("Diagrama AFM\n(Kuno 1968)", fontsize=10, pad=10)

    # Linea divisoria tholeiitic/calc-alkaline (aprox Irvine & Baragar 1971)
    div_a = [0.00,0.05,0.10,0.15,0.20,0.25,0.30]
    div_f = [0.60,0.62,0.65,0.67,0.65,0.58,0.45]
    div_m = [1-a-f for a, f in zip(div_a, div_f)]
    xs = []; ys = []
    for a, f, m in zip(div_a, div_f, div_m):
        x, y = _ternary_coords(a, m, f)
        if x is not None:
            xs.append(x); ys.append(y)
    if xs:
        ax.plot(xs, ys, 'k--', linewidth=1, label='Limite calc-alcalino/thol.')

    ax.text(0.2, 0.5, 'Calc-\nalcalina', fontsize=8, color='#2a5', style='italic')
    ax.text(0.6, 0.25, 'Tholeiítica', fontsize=8, color='#a22', style='italic')

    for r in data:
        a_val = _safe(r.get("Na2O")) + _safe(r.get("K2O"))
        f_val = _safe(r.get("Fe2O3(T)")) * 0.8998  # convierte a FeO equiv
        m_val = _safe(r.get("MgO"))
        x, y = _ternary_coords(a_val, m_val, f_val)
        if x is not None:
            ax.scatter(x, y, s=60, color='#e63946', zorder=5,
                       edgecolors='white', linewidths=0.7)
            ax.annotate(r["UNINORTE_CODE"], (x, y), fontsize=6,
                        xytext=(3, 2), textcoords='offset points')

    ax.legend(fontsize=7, loc='upper right')
    fname = out_dir / "igneas_afm.png"
    fig.tight_layout()
    fig.savefig(fname, dpi=130, bbox_inches='tight')
    plt.close(fig)
    print(f"    -> {fname.name}")
    return f"imagenes/diagramas/igneas_afm.png"


# ---- Harker ----
def _diagrama_harker(data, out_dir):
    fig, axes = plt.subplots(1, 2, figsize=(11, 5))

    pairs = [("MgO", "MgO (%)"), ("Al2O3", "Al₂O₃ (%)")]
    for ax, (col, ylabel) in zip(axes, pairs):
        for r in data:
            sio2 = _safe(r.get("SiO2"))
            y    = _safe(r.get(col))
            if sio2 > 0:
                ax.scatter(sio2, y, s=60, color='#e63946', zorder=5,
                           edgecolors='white', linewidths=0.7)
                ax.annotate(r["UNINORTE_CODE"], (sio2, y), fontsize=6,
                            xytext=(3, 2), textcoords='offset points')
        ax.set_xlabel("SiO₂ (%)", fontsize=10)
        ax.set_ylabel(ylabel, fontsize=10)
        ax.set_title(f"Harker: SiO₂ vs {ylabel.split('(')[0].strip()}", fontsize=10)
        ax.grid(True, alpha=0.3, linestyle=':')

    fig.suptitle("Diagramas de Variacion de Harker", fontsize=11, fontweight='bold')
    fname = out_dir / "igneas_harker.png"
    fig.tight_layout()
    fig.savefig(fname, dpi=130, bbox_inches='tight')
    plt.close(fig)
    print(f"    -> {fname.name}")
    return f"imagenes/diagramas/igneas_harker.png"


# ---- Shervais V vs Ti ----
def _diagrama_shervais(data, out_dir):
    fig, ax = plt.subplots(figsize=(7, 5.5))

    # Campos de Shervais (1982) — V vs Ti (ppm)
    campos_shervais = [
        ("MORB", [[0,0],[8000,50],[8000,0]], '#aaddff', 0.4),
        ("IAT",  [[0,0],[0,50],[4000,280],[8000,50]], '#ffe0a0', 0.4),
        ("Boninita", [[0,0],[0,100],[2000,200],[4000,280],[0,50]], '#ffc0c0', 0.3),
        ("Alcalino", [[0,100],[0,500],[8000,500],[8000,280],[4000,280]], '#d0ffd0', 0.3),
    ]
    for nombre, pts, color, alpha in campos_shervais:
        poly = MplPolygon(pts, closed=True, fc=color, ec='gray',
                          alpha=alpha, linewidth=0.7)
        ax.add_patch(poly)
        xs = [p[0] for p in pts]
        ys = [p[1] for p in pts]
        ax.text(sum(xs)/len(xs), sum(ys)/len(ys), nombre,
                ha='center', va='center', fontsize=8, color='#444', style='italic')

    for r in data:
        ti_ppm = _safe(r.get("TiO2")) * 10000 / 79.90 * 47.87  # TiO2% -> Ti ppm aprox
        v      = _safe(r.get("V"))
        if ti_ppm > 0 or v > 0:
            ax.scatter(ti_ppm, v, s=60, color='#e63946', zorder=5,
                       edgecolors='white', linewidths=0.7)
            ax.annotate(r["UNINORTE_CODE"], (ti_ppm, v), fontsize=6,
                        xytext=(3, 2), textcoords='offset points')

    ax.set_xlim(0, 8000)
    ax.set_ylim(0, 500)
    ax.set_xlabel("Ti (ppm)", fontsize=10)
    ax.set_ylabel("V (ppm)", fontsize=10)
    ax.set_title("Diagrama de Shervais (1982)\nV vs Ti", fontsize=10)
    ax.grid(True, alpha=0.3, linestyle=':')

    fname = out_dir / "igneas_shervais.png"
    fig.tight_layout()
    fig.savefig(fname, dpi=130, bbox_inches='tight')
    plt.close(fig)
    print(f"    -> {fname.name}")
    return f"imagenes/diagramas/igneas_shervais.png"


# ---- CIA (Chemical Index of Alteration) ----
def _diagrama_cia(data, out_dir):
    fig, ax = plt.subplots(figsize=(8, 5))
    codes = []; cias = []
    for r in data:
        al  = _mol(r.get("Al2O3"), "Al2O3")
        cao = _mol(r.get("CaO"),   "CaO")
        na  = _mol(r.get("Na2O"),  "Na2O")
        k   = _mol(r.get("K2O"),   "K2O")
        p   = _mol(r.get("P2O5"),  "P2O5")
        cao_corr = max(0, cao - 10/3 * p)
        denom = al + cao_corr + na + k
        if denom > 0:
            cia = al / denom * 100
            codes.append(r["UNINORTE_CODE"])
            cias.append(cia)

    if not codes:
        plt.close(fig)
        return None

    colors = ['#f4b400' if c < 60 else '#e6634a' if c < 80 else '#c0392b' for c in cias]
    bars = ax.bar(codes, cias, color=colors, edgecolor='white', linewidth=0.7)

    # Bandas de referencia
    ax.axhspan(50, 65, alpha=0.08, color='green', label='Rocas frescas (50-65)')
    ax.axhspan(65, 85, alpha=0.08, color='orange', label='Meteorizacion moderada (65-85)')
    ax.axhspan(85, 100, alpha=0.08, color='red', label='Meteorizacion intensa (>85)')

    for bar, cia in zip(bars, cias):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
                f'{cia:.0f}', ha='center', va='bottom', fontsize=7)

    ax.set_ylim(0, 105)
    ax.set_ylabel("CIA", fontsize=10)
    ax.set_title("Indice de Alteracion Quimica (CIA)\nNesbitt & Young (1982)", fontsize=10)
    ax.tick_params(axis='x', rotation=45, labelsize=8)
    ax.grid(True, axis='y', alpha=0.3, linestyle=':')
    ax.legend(fontsize=7, loc='upper right')

    fname = out_dir / "sed_cia.png"
    fig.tight_layout()
    fig.savefig(fname, dpi=130, bbox_inches='tight')
    plt.close(fig)
    print(f"    -> {fname.name}")
    return f"imagenes/diagramas/sed_cia.png"


# ---- Herron (1988) ----
def _diagrama_herron(data, out_dir):
    fig, ax = plt.subplots(figsize=(7, 6))

    # Lineas de campo de Herron (1988) en log(SiO2/Al2O3) vs log(Fe2O3/K2O)
    # Limites aproximados
    campos_herron = {
        'Arenisca\ncuarzosa':  {'xlim':(-0.1, 0.8), 'ylim':(-2.0,  0.0), 'c':'#ffe0a0'},
        'Subarcosa':            {'xlim':( 0.0, 0.8), 'ylim':( 0.0,  0.6), 'c':'#ffc080'},
        'Arcosa':               {'xlim':(-0.5, 0.0), 'ylim':(-0.5,  0.6), 'c':'#ffa060'},
        'Litarenita\nfeldespat.':{'xlim':(-0.5,0.0), 'ylim':( 0.6,  1.5), 'c':'#ff8060'},
        'Litarenita':           {'xlim':( 0.0, 0.8), 'ylim':( 0.6,  1.8), 'c':'#d08060'},
        'Wacke':                {'xlim':(-0.5, 0.0), 'ylim':( 1.5,  2.5), 'c':'#b0b0ff'},
        'Pelita':               {'xlim':(-0.8,-0.5), 'ylim':(-0.5,  2.5), 'c':'#c0c0e0'},
        'Shale\nFerruginoso':   {'xlim':(-0.1, 0.8), 'ylim':( 1.8,  2.5), 'c':'#e0a0a0'},
    }

    # Lineas divisorias principales
    ax.axhline(y=0.0, color='gray', linewidth=0.8)
    ax.axhline(y=0.6, color='gray', linewidth=0.8)
    ax.axhline(y=1.8, color='gray', linewidth=0.8)
    ax.axvline(x=0.0, color='gray', linewidth=0.8)
    ax.axvline(x=-0.5, color='gray', linewidth=0.8)

    # Etiquetas de campos
    labels_pos = {
        'Arenisca\ncuarzosa':  (0.4, -1.0),
        'Subarcosa':            (0.4,  0.3),
        'Arcosa':               (-0.25, 0.1),
        'Litarenita\nfeldespat.':(-0.25, 1.0),
        'Litarenita':           (0.35,  1.2),
        'Wacke':                (-0.25, 2.0),
        'Pelita':               (-0.65, 1.0),
        'Shale\nFerruginoso':   (0.35,  2.1),
    }
    colors_herron = {
        'Arenisca\ncuarzosa':'#ffe0a0','Subarcosa':'#ffc080','Arcosa':'#ffa060',
        'Litarenita\nfeldespat.':'#ff8060','Litarenita':'#d08060',
        'Wacke':'#b0b0ff','Pelita':'#c0c0e0','Shale\nFerruginoso':'#e0a0a0'
    }
    for nombre, (lx, ly) in labels_pos.items():
        ax.text(lx, ly, nombre, ha='center', va='center', fontsize=7,
                color='#444', style='italic',
                bbox=dict(fc=colors_herron.get(nombre,'white'), alpha=0.5, pad=2, ec='none'))

    for r in data:
        sio2  = _safe(r.get("SiO2"))
        al2o3 = _safe(r.get("Al2O3"))
        fe2o3 = _safe(r.get("Fe2O3(T)"))
        k2o   = _safe(r.get("K2O"))
        if al2o3 > 0 and k2o > 0 and sio2 > 0:
            x = math.log10(sio2 / al2o3)
            y = math.log10(fe2o3 / k2o) if k2o > 0 and fe2o3 > 0 else None
            if y is not None:
                ax.scatter(x, y, s=60, color='#f4b400', zorder=5,
                           edgecolors='white', linewidths=0.7)
                ax.annotate(r["UNINORTE_CODE"], (x, y), fontsize=6,
                            xytext=(3, 2), textcoords='offset points')

    ax.set_xlim(-1.0, 1.0)
    ax.set_ylim(-2.5, 3.0)
    ax.set_xlabel("log(SiO₂/Al₂O₃)", fontsize=10)
    ax.set_ylabel("log(Fe₂O₃/K₂O)", fontsize=10)
    ax.set_title("Diagrama de Herron (1988)\nClasificacion de Rocas Sedimentarias", fontsize=10)
    ax.grid(True, alpha=0.2, linestyle=':')

    fname = out_dir / "sed_herron.png"
    fig.tight_layout()
    fig.savefig(fname, dpi=130, bbox_inches='tight')
    plt.close(fig)
    print(f"    -> {fname.name}")
    return f"imagenes/diagramas/sed_herron.png"


# ---- A-CN-K ternario (Nesbitt & Young 1984) ----
def _diagrama_acnk(data, out_dir):
    fig, ax = plt.subplots(figsize=(6.5, 6))
    _draw_ternary_frame(ax, ['A\n(Al₂O₃)', 'K\n(K₂O)', 'CN\n(CaO*+Na₂O)'])
    ax.set_title("Diagrama A-CN-K (Nesbitt & Young 1984)\nLinea de Meteorizacion", fontsize=9, pad=10)

    for r in data:
        al  = _mol(r.get("Al2O3"), "Al2O3")
        cao = _mol(r.get("CaO"),   "CaO")
        na  = _mol(r.get("Na2O"),  "Na2O")
        k   = _mol(r.get("K2O"),   "K2O")
        p   = _mol(r.get("P2O5"),  "P2O5")
        cao_corr = max(0, cao - 10/3 * p)
        cn = cao_corr + na

        x, y = _ternary_coords(al, k, cn)
        if x is not None:
            ax.scatter(x, y, s=60, color='#f4b400', zorder=5,
                       edgecolors='white', linewidths=0.7)
            ax.annotate(r["UNINORTE_CODE"], (x, y), fontsize=6,
                        xytext=(3, 2), textcoords='offset points')

    # Linea de meteorizacion (de plagioclasa a arcilla)
    weath_pts = [(0.3, 0.3, 0.4), (0.5, 0.1, 0.4), (0.7, 0.1, 0.2), (0.9, 0.05, 0.05)]
    wx, wy = [], []
    for (a, k, cn) in weath_pts:
        x, y = _ternary_coords(a, k, cn)
        if x: wx.append(x); wy.append(y)
    if wx:
        ax.plot(wx, wy, 'k--', linewidth=1, alpha=0.5, label='Tendencia meteorizacion')
        ax.legend(fontsize=7)

    fname = out_dir / "sed_acnk.png"
    fig.tight_layout()
    fig.savefig(fname, dpi=130, bbox_inches='tight')
    plt.close(fig)
    print(f"    -> {fname.name}")
    return f"imagenes/diagramas/sed_acnk.png"


# ---- Winchester & Floyd modificado: Zr/TiO2 vs Y ----
def _diagrama_winchester_floyd(data, out_dir):
    fig, ax = plt.subplots(figsize=(7, 5.5))

    # Campos aproximados en escala log-log
    # Winchester & Floyd (1977) modificado por Pearce (1996)
    campos_wf = [
        ('Riolita/\nDacita',     [0.01, 0.1, 0.1, 0.01],  [10, 10, 200, 200],  '#ffc0c0'),
        ('Andesita',             [0.001,0.01,0.01,0.001],  [10, 10, 200, 200],  '#ffe0a0'),
        ('Basalto\nalcalino',    [0.001,0.01,0.01,0.001],  [2,  2,  10,  10],   '#c0e0ff'),
        ('Basalto\nsubalcalino', [0.0001,0.001,0.001,0.0001],[2, 2, 10,  10],   '#c0ffc0'),
    ]

    for nombre, xs, ys, color in campos_wf:
        pts = list(zip(xs, ys))
        poly = MplPolygon(pts, closed=True, fc=color, ec='gray',
                          alpha=0.4, linewidth=0.7, transform=ax.transData)
        ax.add_patch(poly)

    for r in data:
        zr  = _safe(r.get("Zr"))
        tio2= _safe(r.get("TiO2"))
        y   = _safe(r.get("Y"))
        if tio2 > 0 and zr > 0 and y > 0:
            ratio = zr / (tio2 * 10000 + 1e-9)
            ax.scatter(ratio, y, s=60, color='#2a9d3f', zorder=5,
                       edgecolors='white', linewidths=0.7)
            ax.annotate(r["UNINORTE_CODE"], (ratio, y), fontsize=6,
                        xytext=(3, 2), textcoords='offset points')

    ax.set_xscale('log')
    ax.set_yscale('log')
    ax.set_xlabel("Zr/TiO₂ × 10⁴", fontsize=10)
    ax.set_ylabel("Y (ppm)", fontsize=10)
    ax.set_title("Diagrama Winchester & Floyd (1977)\nZr/TiO₂ vs Y — Rocas Metamorficas/Volcanicas", fontsize=9)
    ax.grid(True, which='both', alpha=0.3, linestyle=':')

    fname = out_dir / "meta_wf.png"
    fig.tight_layout()
    fig.savefig(fname, dpi=130, bbox_inches='tight')
    plt.close(fig)
    print(f"    -> {fname.name}")
    return f"imagenes/diagramas/meta_wf.png"


# ---- Zr vs TiO2 ----
def _diagrama_zr_ti(data, out_dir):
    fig, ax = plt.subplots(figsize=(7, 5.5))

    for r in data:
        zr   = _safe(r.get("Zr"))
        tio2 = _safe(r.get("TiO2"))
        ax.scatter(tio2, zr, s=60, color='#2a9d3f', zorder=5,
                   edgecolors='white', linewidths=0.7)
        ax.annotate(r["UNINORTE_CODE"], (tio2, zr), fontsize=6,
                    xytext=(3, 2), textcoords='offset points')

    ax.set_xlabel("TiO₂ (%)", fontsize=10)
    ax.set_ylabel("Zr (ppm)", fontsize=10)
    ax.set_title("Diagrama Zr vs TiO₂\nRocas Metamorficas", fontsize=10)
    ax.grid(True, alpha=0.3, linestyle=':')

    fname = out_dir / "meta_zrti.png"
    fig.tight_layout()
    fig.savefig(fname, dpi=130, bbox_inches='tight')
    plt.close(fig)
    print(f"    -> {fname.name}")
    return f"imagenes/diagramas/meta_zrti.png"


# ---- ACF ternario ----
def _diagrama_acf(data, out_dir):
    fig, ax = plt.subplots(figsize=(6.5, 6))
    _draw_ternary_frame(ax, ['A\n(Al₂O₃+Fe₂O₃-K₂O)', 'F\n(FeO+MgO+MnO)', 'C\n(CaO-3.33P₂O₅)'])
    ax.set_title("Diagrama ACF\nRocas Metamorficas", fontsize=10, pad=10)

    # Campos de minerales metamorficos
    minerals = [
        ('Granate', 0.3, 0.4, 0.3),
        ('Hornblenda', 0.25, 0.35, 0.4),
        ('Clorita', 0.2, 0.5, 0.3),
        ('Biotita', 0.4, 0.4, 0.2),
    ]
    for nombre, a, f, c in minerals:
        x, y = _ternary_coords(a, f, c)
        if x:
            ax.scatter(x, y, s=80, marker='s', color='#ddd', edgecolors='gray', zorder=3)
            ax.text(x, y-0.05, nombre, ha='center', va='top', fontsize=7,
                    color='#555', style='italic')

    for r in data:
        al   = _mol(r.get("Al2O3"), "Al2O3")
        fe3  = _mol(r.get("Fe2O3(T)"), "Fe2O3") * 0.2
        k    = _mol(r.get("K2O"),   "K2O")
        cao  = _mol(r.get("CaO"),   "CaO")
        p    = _mol(r.get("P2O5"),  "P2O5")
        fe_t = _mol(r.get("Fe2O3(T)"), "Fe2O3") * 0.8998
        mg   = _mol(r.get("MgO"),   "MgO")
        mn   = _mol(r.get("MnO"),   "MnO")

        A = al + fe3 - k
        C = max(0, cao - 3.33 * p)
        F = fe_t + mg + mn

        x, y = _ternary_coords(A, F, C)
        if x is not None:
            ax.scatter(x, y, s=60, color='#2a9d3f', zorder=5,
                       edgecolors='white', linewidths=0.7)
            ax.annotate(r["UNINORTE_CODE"], (x, y), fontsize=6,
                        xytext=(3, 2), textcoords='offset points')

    fname = out_dir / "meta_acf.png"
    fig.tight_layout()
    fig.savefig(fname, dpi=130, bbox_inches='tight')
    plt.close(fig)
    print(f"    -> {fname.name}")
    return f"imagenes/diagramas/meta_acf.png"


# ---------------------------------------------------------------------------
# Generacion del HTML
# ---------------------------------------------------------------------------
def generar_html(filas, drx_data, geochem_data, img_muestra, img_drx,
                 img_laminas, diagramas, ruta_salida: str):
    rows_json    = json.dumps(filas, ensure_ascii=False)
    drx_json     = json.dumps(drx_data, ensure_ascii=False)
    geo_json     = json.dumps(geochem_data, ensure_ascii=False)
    img_m_json   = json.dumps(img_muestra, ensure_ascii=False)
    img_drx_json = json.dumps(img_drx, ensure_ascii=False)
    img_lam_json = json.dumps(img_laminas, ensure_ascii=False)
    diag_json    = json.dumps(diagramas, ensure_ascii=False)

    html = (HTML_TEMPLATE
            .replace("__ROWS_JSON__", rows_json)
            .replace("__DRX_JSON__", drx_json)
            .replace("__GEO_JSON__", geo_json)
            .replace("__IMG_MUESTRA_JSON__", img_m_json)
            .replace("__IMG_DRX_JSON__", img_drx_json)
            .replace("__IMG_LAMINAS_JSON__", img_lam_json)
            .replace("__DIAGRAMAS_JSON__", diag_json))

    Path(ruta_salida).write_text(html, encoding="utf-8")
    print(f"Listo: {len(filas)} muestras -> {ruta_salida}")


# ---------------------------------------------------------------------------
# HTML TEMPLATE
# ---------------------------------------------------------------------------
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Muestras Geologicas — UNINORTE</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.css"/>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.Default.css"/>
<style>
:root{--rojo:#e63946;--amarillo:#f4b400;--verde:#2a9d3f;--gris:#8a8a8a;--azul:#1d3557;}
*{box-sizing:border-box;}
html,body{margin:0;padding:0;height:100%;font-family:'Segoe UI',Arial,sans-serif;color:#222;overflow:hidden;}

/* ---- NAV PRINCIPAL ---- */
#mainNav{position:fixed;top:0;left:0;right:0;z-index:2500;
  background:linear-gradient(135deg,#1d3557,#264653);color:#fff;
  display:flex;align-items:center;justify-content:space-between;
  padding:0 16px;height:50px;box-shadow:0 2px 8px rgba(0,0,0,.3);flex-wrap:nowrap;gap:8px;}
#mainNav h1{font-size:.9rem;margin:0;font-weight:600;white-space:nowrap;flex:1;}
.nav-tabs{display:flex;gap:0;flex-shrink:0;}
.nav-tab{padding:14px 18px;border:none;background:rgba(255,255,255,.1);color:#fff;
  cursor:pointer;font-size:.83rem;font-weight:600;border-bottom:3px solid transparent;
  transition:all .2s;height:50px;}
.nav-tab:hover{background:rgba(255,255,255,.2);}
.nav-tab.active{background:rgba(255,255,255,.15);border-bottom-color:#fff;}
#counter{font-size:.75rem;background:rgba(255,255,255,.1);padding:4px 10px;
  border-radius:8px;white-space:nowrap;flex-shrink:0;}

/* ---- PAGINAS ---- */
.page{position:fixed;top:50px;left:0;right:0;bottom:0;display:none;overflow:hidden;}
.page.active{display:flex;}

/* ---- MAPA ---- */
#page-map{flex-direction:column;}
#map{flex:1;width:100%;}

/* ---- TOGGLE FILTROS ---- */
#toggleFilters{position:absolute;top:10px;left:10px;z-index:1100;
  background:#fff;border:none;border-radius:8px;padding:8px 12px;cursor:pointer;
  box-shadow:0 2px 6px rgba(0,0,0,.3);font-weight:600;font-size:.85rem;}

/* ---- PANEL FILTROS (mapa) ---- */
#filterPanel{position:absolute;top:54px;left:10px;z-index:1100;width:280px;
  max-height:calc(100% - 80px);overflow-y:auto;background:#fff;border-radius:10px;
  padding:14px;box-shadow:0 4px 14px rgba(0,0,0,.25);font-size:.82rem;}
#filterPanel.hidden{display:none;}
#filterPanel h3{margin:0 0 8px;font-size:.9rem;color:#1d3557;
  border-bottom:2px solid #e0e0e0;padding-bottom:6px;}
#filterPanel label{display:block;margin:8px 0 3px;font-weight:600;color:#333;}
#filterPanel select,#filterPanel input[type=text]{
  width:100%;padding:6px 8px;border:1px solid #ccc;border-radius:6px;font-size:.82rem;}
.chk-row{display:flex;align-items:center;gap:6px;margin-top:6px;}
.chk-row label{margin:0;font-weight:500;}
#resetBtn{margin-top:12px;width:100%;padding:7px;background:#264653;color:#fff;
  border:none;border-radius:6px;cursor:pointer;font-weight:600;}
#resetBtn:hover{background:#1d3557;}

/* ---- LEYENDA ---- */
.legend{background:#fff;padding:10px 12px;border-radius:8px;
  box-shadow:0 2px 8px rgba(0,0,0,.3);font-size:.78rem;line-height:1.5;max-width:230px;}
.legend h4{margin:0 0 6px;font-size:.85rem;color:#1d3557;}
.legend .item{display:flex;align-items:center;gap:6px;margin:2px 0;}
.legend .dot{width:14px;height:14px;border-radius:50%;display:inline-block;
  border:2px solid #333;flex-shrink:0;}
.legend .dot.approx{border:2px dashed #000;}

/* ---- POPUP ---- */
.popup-content{font-size:.82rem;line-height:1.5;max-width:270px;}
.popup-content table{width:100%;border-collapse:collapse;}
.popup-content td{padding:2px 4px;vertical-align:top;}
.popup-content td.k{font-weight:600;color:#1d3557;white-space:nowrap;}
.popup-content .warn{background:#fff3cd;border:1px solid #ffe69c;color:#7a5b00;
  padding:5px 7px;border-radius:6px;margin-top:6px;font-size:.78rem;}
.popup-content .ok{background:#d4edda;border:1px solid #b7dfc0;color:#155724;
  padding:5px 7px;border-radius:6px;margin-top:6px;font-size:.78rem;}
.popup-content .analysis{margin-top:5px;font-size:.78rem;color:#0a6b2c;font-weight:600;}
.btn-detalle{display:block;margin-top:10px;padding:7px 12px;background:#1d3557;
  color:#fff;border:none;border-radius:8px;cursor:pointer;font-size:.82rem;
  font-weight:600;width:100%;text-align:center;}
.btn-detalle:hover{background:#264653;}

/* ---- MODAL FICHA ---- */
#modalOverlay{display:none;position:fixed;inset:0;z-index:3000;
  background:rgba(0,0,0,.65);align-items:center;justify-content:center;}
#modalOverlay.open{display:flex;}
#modal{background:#fff;border-radius:12px;width:92vw;max-width:860px;
  max-height:92vh;display:flex;flex-direction:column;overflow:hidden;
  box-shadow:0 8px 32px rgba(0,0,0,.4);}
#modalHeader{background:linear-gradient(135deg,#1d3557,#264653);color:#fff;
  padding:14px 18px;display:flex;justify-content:space-between;align-items:flex-start;gap:10px;flex-shrink:0;}
#modalTitle{font-size:1rem;font-weight:700;line-height:1.3;}
#modalSubtitle{font-size:.8rem;opacity:.85;margin-top:3px;}
.modal-analyses{display:flex;flex-wrap:wrap;gap:6px;margin-top:6px;}
.ma-ok{font-size:.72rem;padding:2px 8px;border-radius:10px;background:#d4edda;color:#155724;font-weight:600;}
.ma-no{font-size:.72rem;padding:2px 8px;border-radius:10px;background:#f8d7da;color:#721c24;font-weight:600;}
#closeModal{background:rgba(255,255,255,.2);border:none;color:#fff;
  border-radius:50%;width:32px;height:32px;font-size:1.1rem;cursor:pointer;
  flex-shrink:0;display:flex;align-items:center;justify-content:center;}
#closeModal:hover{background:rgba(255,255,255,.35);}
.tab-bar{display:flex;background:#f0f0f0;border-bottom:2px solid #ddd;flex-shrink:0;flex-wrap:wrap;}
.tab-btn{padding:10px 16px;border:none;background:none;cursor:pointer;
  font-size:.8rem;font-weight:600;color:#555;border-bottom:3px solid transparent;
  margin-bottom:-2px;transition:all .2s;}
.tab-btn.active{color:#1d3557;border-bottom-color:#1d3557;background:#fff;}
.tab-btn:hover:not(.active){background:#e8e8e8;}
.tab-content{display:none;overflow-y:auto;padding:18px;flex:1;}
.tab-content.active{display:block;}

/* ---- GALERIA ---- */
.gallery{position:relative;background:#111;border-radius:8px;overflow:hidden;
  max-height:380px;display:flex;align-items:center;justify-content:center;}
.gallery img{max-width:100%;max-height:380px;object-fit:contain;display:block;}
.gallery-nav{position:absolute;top:50%;transform:translateY(-50%);
  background:rgba(0,0,0,.5);color:#fff;border:none;border-radius:50%;
  width:36px;height:36px;font-size:1.1rem;cursor:pointer;
  display:flex;align-items:center;justify-content:center;}
.gallery-nav:hover{background:rgba(0,0,0,.75);}
#galleryPrev{left:8px;} #galleryNext{right:8px;}
.gallery-counter{text-align:center;font-size:.78rem;color:#777;margin-top:6px;}
.gallery-thumbs{display:flex;gap:6px;flex-wrap:wrap;margin-top:10px;}
.gallery-thumbs img{width:58px;height:58px;object-fit:cover;border-radius:5px;
  cursor:pointer;opacity:.6;border:2px solid transparent;transition:all .15s;}
.gallery-thumbs img.active{opacity:1;border-color:#1d3557;}
.no-photos{text-align:center;padding:40px 20px;color:#888;font-size:.9rem;}

/* ---- LAMINAS DELGADAS ---- */
.lamina-pair{display:flex;gap:10px;margin-bottom:14px;align-items:flex-start;}
.lamina-pair .lam-col{flex:1;text-align:center;}
.lamina-pair .lam-col img{width:100%;border-radius:6px;border:1px solid #ddd;
  max-height:260px;object-fit:contain;}
.lamina-pair .lam-label{font-size:.75rem;font-weight:600;color:#555;margin-top:4px;}
.lamina-nav{display:flex;align-items:center;gap:10px;margin-bottom:12px;}
.lamina-nav button{padding:6px 14px;border:1px solid #ccc;background:#fff;
  border-radius:6px;cursor:pointer;font-size:.82rem;}
.lamina-nav button:hover{background:#f0f0f0;}
.lamina-counter{font-size:.8rem;color:#666;}

/* ---- DRX PANEL ---- */
.drx-img{text-align:center;margin-bottom:16px;}
.drx-img img{max-width:100%;border-radius:8px;border:1px solid #ddd;}
.drx-img-placeholder{background:#f5f5f5;border:2px dashed #ccc;border-radius:8px;
  padding:30px;text-align:center;color:#999;font-size:.85rem;}
.mineral-table{width:100%;border-collapse:collapse;font-size:.8rem;margin:12px 0;}
.mineral-table th{background:#1d3557;color:#fff;padding:6px 8px;text-align:left;}
.mineral-table td{padding:5px 8px;border-bottom:1px solid #eee;vertical-align:top;}
.mineral-table tr:nth-child(even) td{background:#f8f8f8;}
.section-title{font-size:.87rem;font-weight:700;color:#1d3557;margin:16px 0 6px;
  padding-bottom:4px;border-bottom:2px solid #e0e0e0;}
.description-box{background:#f9f9f9;border-left:3px solid #264653;padding:10px 12px;
  border-radius:0 6px 6px 0;font-size:.8rem;line-height:1.6;color:#333;}
.interpretacion-box{background:#fffbf0;border-left:3px solid #f4b400;padding:10px 12px;
  border-radius:0 6px 6px 0;font-size:.8rem;line-height:1.6;color:#333;}
.nota-box{background:#e8f5e9;border-left:3px solid #2a9d3f;padding:10px 12px;
  border-radius:0 6px 6px 0;font-size:.8rem;line-height:1.6;color:#333;margin-top:12px;}
.ia-badge{display:inline-block;background:#f4b400;color:#333;font-size:.7rem;
  font-weight:700;padding:1px 6px;border-radius:10px;margin-left:6px;vertical-align:middle;}
.no-drx{text-align:center;padding:30px;color:#888;font-size:.88rem;}

/* ---- PANEL NO UBICADAS ---- */
#unlocatedPanel{position:absolute;bottom:10px;left:10px;z-index:1100;
  max-width:320px;max-height:200px;overflow-y:auto;background:#fff;border-radius:10px;
  padding:10px 12px;box-shadow:0 4px 14px rgba(0,0,0,.25);font-size:.78rem;display:none;}
#unlocatedPanel.show{display:block;}
#unlocatedPanel h4{margin:0 0 6px;color:#a4161a;font-size:.85rem;}
#unlocatedPanel ul{margin:0;padding-left:16px;}
#loadingBar{position:absolute;bottom:0;left:0;right:0;z-index:1300;
  background:#1d3557;color:#fff;font-size:.78rem;padding:6px 12px;display:none;}

/* popup analyses */
.popup-analyses{display:flex;flex-direction:column;gap:3px;margin-top:8px;}
.analysis-no{font-size:.78rem;color:#c0392b;font-weight:600;padding:3px 6px;
  background:#fdecea;border-radius:4px;}
.desc-section{margin-bottom:16px;}
.desc-section h4{margin:0 0 8px;font-size:.9rem;color:#1d3557;border-bottom:1px solid #e0e0e0;padding-bottom:4px;}

/* ---- PAGINA DRX Y GEOQUIMICA ---- */
#page-geochem{flex-direction:row;overflow:hidden;}
/* tarjeta DRX+geo por muestra */
.drx-geo-card{background:#fff;border:1px solid #e0e0e0;border-radius:10px;
  padding:16px;margin-bottom:18px;}
.drx-geo-card-header{display:flex;align-items:center;gap:10px;margin-bottom:12px;flex-wrap:wrap;}
.drx-geo-code{font-size:1rem;font-weight:700;color:#1d3557;}
.drx-geo-tipo{font-size:.78rem;color:#888;background:#f0f4f8;padding:2px 8px;border-radius:12px;}
.drx-geo-body{display:flex;gap:16px;flex-wrap:wrap;}
.drx-geo-img{flex:0 0 220px;max-width:220px;}
.drx-geo-img img{width:100%;border-radius:6px;cursor:pointer;}
.drx-geo-img-placeholder{width:100%;height:160px;background:#f5f5f5;border:1px dashed #ccc;
  border-radius:6px;display:flex;align-items:center;justify-content:center;
  font-size:.75rem;color:#aaa;text-align:center;}
.drx-geo-info{flex:1;min-width:200px;}
.drx-geo-popup-btn{background:#1d3557;color:#fff;border:none;border-radius:6px;
  padding:6px 12px;cursor:pointer;font-size:.78rem;font-weight:600;margin-bottom:10px;}
.drx-geo-popup-btn:hover{background:#264653;}
.geo-table-visual{width:100%;border-collapse:collapse;font-size:.78rem;}
.geo-table-visual th{background:#1d3557;color:#fff;padding:5px 8px;text-align:left;}
.geo-table-visual td{padding:4px 8px;border-bottom:1px solid #f0f0f0;}
.geo-table-visual tr:nth-child(even) td{background:#f7f9fc;}
.geo-bar-cell{display:flex;align-items:center;gap:6px;}
.geo-bar{height:10px;background:#e63946;border-radius:3px;min-width:2px;}
/* popup mineral info */
.mineral-popup-overlay{position:fixed;inset:0;background:rgba(0,0,0,.5);z-index:9000;
  display:flex;align-items:center;justify-content:center;}
.mineral-popup{background:#fff;border-radius:10px;padding:20px;max-width:480px;width:90%;
  max-height:80vh;overflow-y:auto;position:relative;}
.mineral-popup h3{margin:0 0 12px;color:#1d3557;}
.mineral-popup-close{position:absolute;top:10px;right:12px;background:none;border:none;
  font-size:1.2rem;cursor:pointer;color:#555;}

/* ---- PAGINA LAMINAS DELGADAS ---- */
#page-laminas{overflow-y:auto;padding:20px;}
.lam-page-sample{background:#fff;border:1px solid #ddd;border-radius:10px;
  margin-bottom:24px;overflow:hidden;}
.lam-page-header{background:#1d3557;color:#fff;padding:12px 16px;display:flex;align-items:center;gap:12px;}
.lam-page-code{font-size:1rem;font-weight:700;}
.lam-page-tipo{font-size:.78rem;opacity:.8;}
.lam-page-body{padding:16px;}
.lam-desc-grid{display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-bottom:16px;}
.lam-desc-box{background:#f7f9fc;border:1px solid #e0e0e0;border-radius:6px;padding:10px;}
.lam-desc-box label{font-size:.72rem;font-weight:700;color:#1d3557;text-transform:uppercase;
  display:block;margin-bottom:4px;}
.lam-desc-box p{margin:0;font-size:.8rem;color:#555;min-height:40px;}
.lam-pair-grid{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:12px;}
.lam-pair-item{flex:1;min-width:200px;max-width:320px;}
.lam-pair-item img{width:100%;border-radius:6px;cursor:pointer;}
.lam-pair-label{font-size:.75rem;font-weight:600;color:#444;margin-top:4px;text-align:center;}
.lam-pair-coords{font-size:.72rem;color:#888;text-align:center;}
.lam-pair-title{font-size:.83rem;font-weight:700;color:#264653;margin:10px 0 6px;
  border-bottom:1px solid #eee;padding-bottom:4px;}
#gc-sidebar{width:260px;min-width:180px;background:#f7f9fc;border-right:1px solid #ddd;
  display:flex;flex-direction:column;overflow:hidden;flex-shrink:0;}
#gc-sidebar-header{padding:12px;background:#1d3557;color:#fff;font-weight:700;
  font-size:.88rem;flex-shrink:0;}
#gc-sidebar-filters{padding:10px;flex-shrink:0;border-bottom:1px solid #ddd;}
#gc-sidebar-filters select,#gc-sidebar-filters input{
  width:100%;padding:5px 7px;border:1px solid #ccc;border-radius:5px;
  font-size:.8rem;margin-top:4px;}
#gc-sample-list{flex:1;overflow-y:auto;padding:4px 0;}
.gc-sample-item{padding:8px 12px;cursor:pointer;border-bottom:1px solid #eee;
  font-size:.8rem;transition:background .15s;}
.gc-sample-item:hover{background:#e8f0fe;}
.gc-sample-item.selected{background:#d0e4ff;font-weight:600;}
.gc-sample-item .gc-code{font-weight:700;color:#1d3557;}
.gc-sample-item .gc-tipo{font-size:.72rem;color:#888;}
#gc-main{flex:1;display:flex;flex-direction:column;overflow:hidden;}
#gc-rock-tabs{display:flex;background:#fff;border-bottom:2px solid #e0e0e0;
  flex-shrink:0;padding:0 12px;}
.gc-rock-tab{padding:10px 20px;border:none;background:none;cursor:pointer;
  font-size:.84rem;font-weight:600;color:#666;border-bottom:3px solid transparent;
  margin-bottom:-2px;transition:all .2s;}
.gc-rock-tab.active{color:#1d3557;border-bottom-color:#1d3557;}
.gc-rock-tab:hover:not(.active){background:#f0f0f0;}
#gc-content{flex:1;overflow-y:auto;padding:16px;display:flex;flex-wrap:wrap;gap:14px;}
.gc-diagram-card{background:#fff;border:1px solid #e0e0e0;border-radius:10px;
  overflow:hidden;box-shadow:0 2px 6px rgba(0,0,0,.08);}
.gc-diagram-card img{width:100%;display:block;}
.gc-diagram-title{padding:6px 10px;font-size:.78rem;font-weight:600;color:#555;
  background:#f8f8f8;border-top:1px solid #eee;}
.gc-no-data{padding:40px;text-align:center;color:#888;font-size:.9rem;width:100%;}
#gc-selected-card{position:fixed;bottom:16px;right:16px;z-index:1200;
  background:#fff;border-radius:12px;box-shadow:0 4px 20px rgba(0,0,0,.25);
  padding:14px;max-width:340px;font-size:.8rem;display:none;}
#gc-selected-card.show{display:block;}
#gc-selected-card .gc-card-close{float:right;cursor:pointer;color:#888;font-size:1rem;}
#gc-selected-card h4{margin:0 0 8px;font-size:.88rem;color:#1d3557;}
.gc-values-table{width:100%;border-collapse:collapse;margin-top:6px;}
.gc-values-table td{padding:3px 6px;border-bottom:1px solid #f0f0f0;font-size:.77rem;}
.gc-values-table td:first-child{font-weight:600;color:#444;width:50%;}
.btn-ver-ficha{margin-top:8px;width:100%;padding:7px;background:#1d3557;color:#fff;
  border:none;border-radius:7px;cursor:pointer;font-weight:600;font-size:.82rem;}
.btn-ver-ficha:hover{background:#264653;}

@media(max-width:700px){
  #gc-sidebar{width:200px;}
  #mainNav h1{display:none;}
  .nav-tab{padding:14px 12px;font-size:.78rem;}
  #modal{width:98vw;max-height:96vh;}
  .lamina-pair{flex-direction:column;}
}
</style>
</head>
<body>

<!-- NAVEGACION PRINCIPAL -->
<nav id="mainNav">
  <h1>Muestras Geologicas — Petrografia / XRD UNINORTE</h1>
  <div class="nav-tabs">
    <button class="nav-tab active" onclick="showPage('map')">🗺 Mapa</button>
    <button class="nav-tab" onclick="showPage('geochem')">🔬 DRX y geoquímica</button>
    <button class="nav-tab" onclick="showPage('laminas')">🪨 Láminas delgadas</button>
  </div>
  <div id="counter">Cargando...</div>
</nav>

<!-- PAGINA: MAPA -->
<div id="page-map" class="page active">
  <button id="toggleFilters">&#9776; Filtros</button>
  <div id="filterPanel">
    <h3>Filtros</h3>
    <label for="searchCode">Buscar por UNINORTE CODE</label>
    <input type="text" id="searchCode" placeholder="Ej: NYB05">
    <label for="fPais">Pais</label>
    <select id="fPais"><option value="">Todos</option></select>
    <label for="fTipo">Tipo de roca</label>
    <select id="fTipo"><option value="">Todos</option></select>
    <label for="fAutor">Autor</label>
    <select id="fAutor"><option value="">Todos</option></select>
    <label for="fSpp">Especie (spp)</label>
    <select id="fSpp"><option value="">Todas</option></select>
    <div class="chk-row"><input type="checkbox" id="fDRX"><label for="fDRX">Solo con analisis DRX</label></div>
    <div class="chk-row"><input type="checkbox" id="fGeo"><label for="fGeo">Solo con analisis Geoquimico</label></div>
    <div class="chk-row"><input type="checkbox" id="fPetro"><label for="fPetro">Solo analisis petrografico</label></div>
    <button id="resetBtn">Limpiar filtros</button>
  </div>
  <div id="unlocatedPanel"></div>
  <div id="loadingBar"></div>
  <div id="map"></div>
</div>

<!-- PAGINA: DRX Y GEOQUIMICA -->
<div id="page-geochem" class="page">
  <div id="gc-sidebar">
    <div id="gc-sidebar-header">Muestras</div>
    <div id="gc-sidebar-filters">
      <select id="gc-filter-tipo">
        <option value="">Todos los tipos</option>
        <option value="IGNEA">Ígnea</option>
        <option value="SEDIMENTARIA">Sedimentaria</option>
        <option value="METAMORFICA">Metamórfica</option>
      </select>
    </div>
    <div id="gc-sample-list"></div>
  </div>
  <div id="gc-main">
    <div id="gc-content"></div>
  </div>
</div>

<!-- PAGINA: LAMINAS DELGADAS -->
<div id="page-laminas" class="page">
  <div id="lam-page-content"></div>
</div>

<!-- MODAL FICHA COMPLETA -->
<div id="modalOverlay">
  <div id="modal">
    <div id="modalHeader">
      <div>
        <div id="modalTitle">—</div>
        <div id="modalSubtitle">—</div>
        <div id="modalAnalyses" class="modal-analyses"></div>
      </div>
      <button id="closeModal">&#x2715;</button>
    </div>
    <div class="tab-bar">
      <button class="tab-btn active" data-tab="tabDesc">📋 Descripción</button>
      <button class="tab-btn" data-tab="tabMuestra">📷 Fotos muestra macroscópica</button>
      <button class="tab-btn" data-tab="tabLaminas">🔬 Láminas delgadas</button>
      <button class="tab-btn" data-tab="tabDRX">📊 Análisis DRX EVA</button>
    </div>
    <div id="tabDesc" class="tab-content active">
      <div id="descBody"></div>
    </div>
    <div id="tabMuestra" class="tab-content">
      <div class="gallery">
        <img id="galleryImg" src="" alt="Foto muestra">
        <button class="gallery-nav" id="galleryPrev">&#8249;</button>
        <button class="gallery-nav" id="galleryNext">&#8250;</button>
      </div>
      <div class="gallery-counter" id="galleryCounter"></div>
      <div class="gallery-thumbs" id="galleryThumbs"></div>
    </div>
    <div id="tabLaminas" class="tab-content">
      <div id="laminasBody"></div>
    </div>
    <div id="tabDRX" class="tab-content">
      <div class="drx-img" id="drxImgWrap"></div>
      <div id="drxBody"><div class="no-drx">No hay datos DRX EVA para esta muestra.</div></div>
    </div>
  </div>
</div>

<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet.markercluster@1.5.3/dist/leaflet.markercluster.js"></script>
<script>
/* ============================================================
   DATOS
   ============================================================ */
const HEADER = ["UNINORTE_CODE","DESTRUCTIVE","GARDEN_CODE","PAIS","UBICACION",
  "SPP","FECHA","LATITUD","LONGITUD","AUTOR","TIPO_ROCA","CLASIFICACION",
  "PESO_GEOQ","OBS_PETRO","GEOQUIMICA","DRX","PESO_TOTAL"];

const RAW_ROWS    = __ROWS_JSON__;
const DRX_DATA    = __DRX_JSON__;
const GEO_DATA    = __GEO_JSON__;  // array de objetos {UNINORTE_CODE, SiO2, ...}
const IMG_MUESTRA = __IMG_MUESTRA_JSON__;
const IMG_DRX     = __IMG_DRX_JSON__;
const IMG_LAMINAS = __IMG_LAMINAS_JSON__;
const DIAGRAMAS   = __DIAGRAMAS_JSON__;

const samples = RAW_ROWS.map(r => {
  const o = {};
  HEADER.forEach((h,i) => o[h] = r[i]);
  return o;
});

// Mapa rapido por codigo
const sampleByCode = {};
samples.forEach(s => sampleByCode[s.UNINORTE_CODE] = s);

const geoByCode = {};
GEO_DATA.forEach(g => geoByCode[g.UNINORTE_CODE] = g);

/* helpers */
function cleanAuthor(a){return a ? String(a).replace(/[0-9]/g,"").replace(/\s{2,}/g," ").trim() : "";}
function nonEmpty(v){return v !== null && v !== undefined && String(v).trim() !== "";}
function parseCoord(v){if(!nonEmpty(v)) return null; const n=parseFloat(String(v).replace(",",".")); return isNaN(n)?null:n;}
function esc(s){return s ? String(s).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;") : "—";}

/* ============================================================
   NAVEGACION PRINCIPAL
   ============================================================ */
function showPage(name){
  document.querySelectorAll('.page').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.nav-tab').forEach(t => t.classList.remove('active'));
  document.getElementById('page-'+name).classList.add('active');
  document.querySelectorAll('.nav-tab').forEach(t => {
    if(t.getAttribute('onclick') && t.getAttribute('onclick').includes("'"+name+"'"))
      t.classList.add('active');
  });
  if(name === 'map') setTimeout(()=>map.invalidateSize(), 50);
  if(name === 'geochem') renderGcPage();
  if(name === 'laminas') renderLaminasPage();
}

/* ============================================================
   MAPA
   ============================================================ */
const ROCK_COLORS = {IGNEA:"#e63946",SEDIMENTARIA:"#f4b400",METAMORFICA:"#2a9d3f"};
const DEFAULT_COLOR = "#8a8a8a";
function colorForRock(t){return ROCK_COLORS[String(t||"").toUpperCase().trim()] || DEFAULT_COLOR;}
function makeIcon(color,approx){
  const b = approx?"3px dashed #000":"2px solid #222";
  const sz = approx?22:18;
  return L.divIcon({className:"",
    html:`<div style="width:${sz}px;height:${sz}px;border-radius:50%;background:${color};border:${b};box-shadow:0 1px 3px rgba(0,0,0,.5)"></div>`,
    iconSize:[sz,sz],iconAnchor:[sz/2,sz/2]});
}

const map = L.map("map",{zoomControl:true}).setView([5,-55],3);
L.tileLayer("https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png",{
  maxZoom:19,attribution:'&copy; <a href="https://www.openstreetmap.org/copyright">OpenStreetMap</a>'
}).addTo(map);
L.control.scale({metric:true,imperial:false}).addTo(map);

const markerCluster = L.markerClusterGroup();
map.addLayer(markerCluster);

const legend = L.control({position:"bottomright"});
legend.onAdd = function(){
  const d = L.DomUtil.create("div","legend");
  d.innerHTML=`<h4>Leyenda</h4>
    <div class="item"><span class="dot" style="background:${ROCK_COLORS.IGNEA}"></span> Ignea</div>
    <div class="item"><span class="dot" style="background:${ROCK_COLORS.SEDIMENTARIA}"></span> Sedimentaria</div>
    <div class="item"><span class="dot" style="background:${ROCK_COLORS.METAMORFICA}"></span> Metamorfica</div>
    <div class="item"><span class="dot" style="background:${DEFAULT_COLOR}"></span> Otro</div>
    <hr style="margin:6px 0;border:none;border-top:1px solid #ddd;">
    <div class="item"><span class="dot" style="background:#999"></span> Coord. originales</div>
    <div class="item"><span class="dot approx" style="background:#999"></span> Ubicacion aprox.</div>`;
  return d;
};
legend.addTo(map);

/* geocoding */
const geocodeCache = {};
async function geocodeQuery(q){
  if(geocodeCache[q]!==undefined) return geocodeCache[q];
  try{
    const r=await fetch(`https://nominatim.openstreetmap.org/search?format=json&limit=1&q=${encodeURIComponent(q)}`,{headers:{"Accept-Language":"es"}});
    if(!r.ok){geocodeCache[q]=null;return null;}
    const data=await r.json();
    if(data&&data.length){const res={lat:parseFloat(data[0].lat),lon:parseFloat(data[0].lon)};geocodeCache[q]=res;return res;}
    geocodeCache[q]=null;return null;
  }catch{geocodeCache[q]=null;return null;}
}
function sleep(ms){return new Promise(r=>setTimeout(r,ms));}
function stripAdminPrefixes(t){return t.replace(/^(MUNICIPIO\s+DE|MUNICIPIO|PROVINCIA\s+DE|DEPARTAMENTO\s+DE|DISTRITO\s+DE|ESTADO\s+DE)\s+/i,"").trim();}
async function geocodeSample(ubicacion,pais){
  const tokensRaw=(ubicacion||"").split(",").map(t=>t.trim()).filter(Boolean);
  const tokens=tokensRaw.map(stripAdminPrefixes).filter(Boolean);
  const intentos=[];
  if(ubicacion&&pais) intentos.push(`${ubicacion}, ${pais}`);
  if(tokens.length>=2&&pais) intentos.push(`${tokens[tokens.length-1]}, ${tokens[0]}, ${pais}`);
  if(tokens.length>=1&&pais){intentos.push(`${tokens[tokens.length-1]}, ${pais}`);if(tokens[0]!==tokens[tokens.length-1]) intentos.push(`${tokens[0]}, ${pais}`);}
  if(pais) intentos.push(pais);
  for(const q of intentos){const r=await geocodeQuery(q);await sleep(1100);if(r) return r;}
  return null;
}

function buildPopup(s,approx){
  const tieneDRX=nonEmpty(s.DRX), tieneGeoq=nonEmpty(s.GEOQUIMICA);
  const tienePetro=nonEmpty(s.OBS_PETRO)||IMG_LAMINAS[s.UNINORTE_CODE]?.length>0;
  let html=`<div class="popup-content"><table>
    <tr><td class="k">UNINORTE CODE</td><td>${esc(s.UNINORTE_CODE)}</td></tr>
    <tr><td class="k">País</td><td>${esc(s.PAIS)}</td></tr>
    <tr><td class="k">Ubicación</td><td>${esc(s.UBICACION)}</td></tr>
    <tr><td class="k">Especie</td><td>${esc(s.SPP)}</td></tr>
    <tr><td class="k">Fecha</td><td>${esc(s.FECHA)}</td></tr>
    <tr><td class="k">Autor</td><td>${esc(s.AUTOR)}</td></tr>
    <tr><td class="k">Tipo de roca</td><td>${esc(s.TIPO_ROCA)}</td></tr>
    <tr><td class="k">Clasificación</td><td>${esc(s.CLASIFICACION)}</td></tr>
    </table>
    <div class="popup-analyses">
      <div class="${tienePetro?'analysis':'analysis-no'}">
        ${tienePetro?'✓':'✗'} Análisis petrográfico</div>
      <div class="${tieneDRX?'analysis':'analysis-no'}">
        ${tieneDRX?'✓':'✗'} Análisis DRX EVA</div>
      <div class="${tieneGeoq?'analysis':'analysis-no'}">
        ${tieneGeoq?'✓':'✗'} Análisis geoquímico</div>
    </div>`;
  if(approx) html+=`<div class="warn">⚠ Ubicación aproximada.</div>`;
  else       html+=`<div class="ok">📍 Coordenadas originales.</div>`;
  html+=`<button class="btn-detalle" onclick="abrirFicha('${esc(s.UNINORTE_CODE)}')">Ver ficha completa →</button></div>`;
  return html;
}

let markerRecords=[];
let countOriginal=0,countGeocoded=0,countUnlocated=0;
const unlocatedList=[];

function updateCounter(){
  const total=markerRecords.length;
  document.getElementById("counter").innerHTML=
    `<b>${total}</b> muestras | orig:<b>${countOriginal}</b> geo:<b>${countGeocoded}</b> sin:<b>${countUnlocated}</b>`;
}
function renderUnlocatedPanel(){
  const p=document.getElementById("unlocatedPanel");
  if(!unlocatedList.length){p.classList.remove("show");return;}
  p.classList.add("show");
  p.innerHTML=`<h4>&#9888; Sin georref. (${unlocatedList.length})</h4>
    <ul>${unlocatedList.map(u=>`<li><b>${u.code}</b> — ${u.pais||"—"}, ${u.ubicacion||"—"}</li>`).join("")}</ul>`;
}
function addSampleMarker(s,lat,lon,approx){
  const color=colorForRock(s.TIPO_ROCA);
  const marker=L.marker([lat,lon],{icon:makeIcon(color,approx)});
  marker.bindPopup(buildPopup(s,approx),{maxWidth:300});
  markerCluster.addLayer(marker);
  markerRecords.push({sample:s,marker,approx,lat,lon});
}

/* Coordenadas de respaldo para muestras sin lat/lon en el Excel */
const COORD_FALLBACK = {
  "NYB01": {lat:  8.50, lon: -82.50},
  "NYB02": {lat:  8.50, lon: -82.50},
  "NYB14": {lat:  4.00, lon: -55.50},
  "NYB18": {lat:  6.50, lon: -64.70},
  "NYB22": {lat:  4.00, lon: -66.00},
  "NYB23": {lat:  4.00, lon: -67.00},
  "NYB24": {lat: -27.50, lon: -55.00},
  "NYB25": {lat: -27.10, lon: -54.70},
  "NYB28": {lat: 10.50, lon: -84.50},
  "NYB37": {lat: -3.10, lon: -58.50}
};

async function init(){
  const loadingBar=document.getElementById("loadingBar");
  const pending=[];
  samples.forEach(s=>{
    const lat=parseCoord(s.LATITUD),lon=parseCoord(s.LONGITUD);
    if(lat!==null&&lon!==null){countOriginal++;addSampleMarker(s,lat,lon,false);}
    else if(COORD_FALLBACK[s.UNINORTE_CODE]){
      const fb=COORD_FALLBACK[s.UNINORTE_CODE];
      countGeocoded++;addSampleMarker(s,fb.lat,fb.lon,true);
    } else pending.push(s);
  });
  updateCounter();
  populateFilterOptions();
  if(pending.length){
    loadingBar.style.display="block";
    for(let i=0;i<pending.length;i++){
      const s=pending[i];
      loadingBar.textContent=`Geocodificando (${i+1}/${pending.length}) — ${s.UNINORTE_CODE}`;
      const r=await geocodeSample(s.UBICACION,s.PAIS);
      if(r){countGeocoded++;addSampleMarker(s,r.lat,r.lon,true);}
      else{countUnlocated++;unlocatedList.push({code:s.UNINORTE_CODE,pais:s.PAIS,ubicacion:s.UBICACION});}
      updateCounter();renderUnlocatedPanel();
    }
    loadingBar.style.display="none";
  }
}

/* filtros */
function uniqueSorted(arr){return[...new Set(arr.filter(v=>nonEmpty(v)))].sort();}
function fillSelect(id,vals){
  const sel=document.getElementById(id);
  vals.forEach(v=>{const o=document.createElement("option");o.value=v;o.textContent=v;sel.appendChild(o);});
}
function populateFilterOptions(){
  fillSelect("fPais",uniqueSorted(samples.map(s=>s.PAIS)));
  fillSelect("fTipo",uniqueSorted(samples.map(s=>s.TIPO_ROCA)));
  fillSelect("fAutor",uniqueSorted(samples.map(s=>cleanAuthor(s.AUTOR))));
  fillSelect("fSpp",uniqueSorted(samples.map(s=>s.SPP)));
}
function applyFilters(){
  const code=document.getElementById("searchCode").value.trim().toUpperCase();
  const pais=document.getElementById("fPais").value;
  const tipo=document.getElementById("fTipo").value;
  const autor=document.getElementById("fAutor").value;
  const spp=document.getElementById("fSpp").value;
  const onlyDRX=document.getElementById("fDRX").checked;
  const onlyGeo=document.getElementById("fGeo").checked;
  const onlyPetro=document.getElementById("fPetro").checked;
  markerCluster.clearLayers();
  markerRecords.forEach(rec=>{
    const s=rec.sample;let v=true;
    if(code && !String(s.UNINORTE_CODE||"").toUpperCase().includes(code)) v=false;
    if(pais && s.PAIS!==pais) v=false;
    if(tipo && s.TIPO_ROCA!==tipo) v=false;
    if(autor && cleanAuthor(s.AUTOR)!==autor) v=false;
    if(spp && s.SPP!==spp) v=false;
    if(onlyDRX && !nonEmpty(s.DRX)) v=false;
    if(onlyGeo && !nonEmpty(s.GEOQUIMICA)) v=false;
    if(onlyPetro && (nonEmpty(s.DRX)||nonEmpty(s.GEOQUIMICA))) v=false;
    if(v) markerCluster.addLayer(rec.marker);
  });
}
["searchCode","fPais","fTipo","fAutor","fSpp"].forEach(id=>{
  document.getElementById(id).addEventListener("input",applyFilters);
  document.getElementById(id).addEventListener("change",applyFilters);
});
["fDRX","fGeo","fPetro"].forEach(id=>document.getElementById(id).addEventListener("change",applyFilters));
document.getElementById("resetBtn").addEventListener("click",()=>{
  document.getElementById("searchCode").value="";
  ["fPais","fTipo","fAutor","fSpp"].forEach(id=>document.getElementById(id).value="");
  ["fDRX","fGeo","fPetro"].forEach(id=>document.getElementById(id).checked=false);
  applyFilters();
});
document.getElementById("toggleFilters").addEventListener("click",()=>
  document.getElementById("filterPanel").classList.toggle("hidden"));
if(window.innerWidth<700) document.getElementById("filterPanel").classList.add("hidden");

/* ============================================================
   MODAL FICHA
   ============================================================ */
let galleryImgs=[];
let galleryIdx=0;
let laminaPairs=[];
let laminaIdx=0;

function abrirFicha(code){
  const s=sampleByCode[code];
  if(!s) return;
  document.getElementById("modalTitle").textContent=
    `${code}  |  ${s.TIPO_ROCA||"Tipo de roca no especificado"}`;
  document.getElementById("modalSubtitle").textContent=
    `${s.UBICACION||"—"}, ${s.PAIS||"—"}  •  ${s.SPP||""}`;

  // Indicadores de análisis
  const tienePetro=nonEmpty(s.OBS_PETRO)||(IMG_LAMINAS[code]||[]).length>0;
  const tieneDRX=nonEmpty(s.DRX);
  const tieneGeoq=nonEmpty(s.GEOQUIMICA);
  document.getElementById("modalAnalyses").innerHTML=
    `<span class="${tienePetro?'ma-ok':'ma-no'}">${tienePetro?'✓':'✗'} Análisis petrográfico</span>` +
    `<span class="${tieneDRX?'ma-ok':'ma-no'}">${tieneDRX?'✓':'✗'} Análisis DRX EVA</span>` +
    `<span class="${tieneGeoq?'ma-ok':'ma-no'}">${tieneGeoq?'✓':'✗'} Análisis geoquímico</span>`;

  // Tab descripción
  renderDesc(code, s);

  // Tab muestra
  galleryImgs=IMG_MUESTRA[code]||[];
  galleryIdx=0;
  renderGallery();

  // Tab laminas
  laminaPairs=IMG_LAMINAS[code]||[];
  laminaIdx=0;
  renderLaminas(s);

  // Tab DRX
  renderDRX(code,s);

  activarTab("tabDesc");
  document.getElementById("modalOverlay").classList.add("open");
}

function renderDesc(code, s){
  const body=document.getElementById("descBody");
  const obs=s.OBS_PETRO||"";
  const d=DRX_DATA[code];
  const g=geoByCode[code];
  let html=`<div class="desc-section"><h4>Descripción petrográfica</h4>`;
  if(obs) html+=`<div class="description-box">${esc(obs)}</div>`;
  else    html+=`<div class="no-drx">Sin descripción petrográfica registrada.</div>`;
  html+=`</div>`;
  if(d&&d.descripcion){
    html+=`<div class="desc-section"><h4>Descripción propia (DRX)</h4>
      <div class="description-box">${esc(d.descripcion)}</div></div>`;
  }
  if(g){
    const rows=SHOW_COLS.filter(c=>g[c]!==undefined&&g[c]!==null).map(c=>
      `<tr><td><b>${c}</b></td><td>${Number(g[c]).toFixed(2)}</td><td>${GEO_UNITS[c]||''}</td></tr>`).join("");
    if(rows) html+=`<div class="desc-section"><h4>Geoquímica</h4>
      <table class="mineral-table"><thead><tr><th>Elemento</th><th>Valor</th><th>Unidad</th></tr></thead>
      <tbody>${rows}</tbody></table></div>`;
  }
  body.innerHTML=html;
}

function renderGallery(){
  const tab=document.getElementById("tabMuestra");
  if(galleryImgs.length===0){
    tab.innerHTML=`<div class="no-photos">&#128247; No hay imagenes de muestra de mano para esta muestra.</div>`;
    return;
  }
  if(!document.getElementById("galleryImg")){
    tab.innerHTML=`<div class="gallery"><img id="galleryImg" src="" alt="">
      <button class="gallery-nav" id="galleryPrev">&#8249;</button>
      <button class="gallery-nav" id="galleryNext">&#8250;</button></div>
      <div class="gallery-counter" id="galleryCounter"></div>
      <div class="gallery-thumbs" id="galleryThumbs"></div>`;
    bindGalleryNav();
  }
  document.getElementById("galleryImg").src=galleryImgs[galleryIdx];
  document.getElementById("galleryCounter").textContent=`${galleryIdx+1} / ${galleryImgs.length}`;
  const thumbs=document.getElementById("galleryThumbs");
  thumbs.innerHTML="";
  galleryImgs.forEach((src,i)=>{
    const t=document.createElement("img");
    t.src=src;t.alt="";
    if(i===galleryIdx) t.classList.add("active");
    t.addEventListener("click",()=>{galleryIdx=i;actualizarGallery();});
    thumbs.appendChild(t);
  });
}
function actualizarGallery(){
  const img=document.getElementById("galleryImg");
  if(!img) return;
  img.src=galleryImgs[galleryIdx];
  document.getElementById("galleryCounter").textContent=`${galleryIdx+1} / ${galleryImgs.length}`;
  document.getElementById("galleryThumbs").querySelectorAll("img").forEach((t,i)=>
    t.classList.toggle("active",i===galleryIdx));
}
function bindGalleryNav(){
  document.getElementById("galleryPrev")?.addEventListener("click",()=>{
    if(!galleryImgs.length) return;
    galleryIdx=(galleryIdx-1+galleryImgs.length)%galleryImgs.length;actualizarGallery();});
  document.getElementById("galleryNext")?.addEventListener("click",()=>{
    if(!galleryImgs.length) return;
    galleryIdx=(galleryIdx+1)%galleryImgs.length;actualizarGallery();});
}
bindGalleryNav();

function renderLaminas(s){
  const body=document.getElementById("laminasBody");
  const tienePetro=nonEmpty(s&&s.OBS_PETRO)||laminaPairs.length>0;
  if(!tienePetro){
    body.innerHTML=`<div class="no-photos">Esta muestra no presenta análisis petrográfico. No hay lámina delgada disponible.</div>`;
    return;
  }
  if(!laminaPairs||laminaPairs.length===0){
    body.innerHTML=`<div class="no-photos">🔬 No hay imágenes de lámina delgada para esta muestra.</div>`;
    return;
  }
  function renderPair(idx){
    const pair=laminaPairs[idx];
    let html=`<div class="lamina-nav">
      <button onclick="lamNav(-1)">&#8249; Anterior</button>
      <span class="lamina-counter">Par ${idx+1} / ${laminaPairs.length}</span>
      <button onclick="lamNav(1)">Siguiente &#8250;</button>
    </div>
    <div class="lamina-pair">
      <div class="lam-col">
        ${pair.nc ? `<img src="${pair.nc}" alt="NC">` : '<div class="no-photos" style="height:200px">Sin imagen NC</div>'}
        <div class="lam-label">NC — Nicoles Cruzados</div>
      </div>
      <div class="lam-col">
        ${pair.np ? `<img src="${pair.np}" alt="NP">` : '<div class="no-photos" style="height:200px">Sin imagen NP</div>'}
        <div class="lam-label">NP — Nicoles Paralelos</div>
      </div>
    </div>`;
    body.innerHTML=html;
  }
  renderPair(laminaIdx);
  window.lamNav=function(dir){
    laminaIdx=(laminaIdx+dir+laminaPairs.length)%laminaPairs.length;
    renderPair(laminaIdx);
  };
}

function renderDRX(code,s){
  const imgWrap=document.getElementById("drxImgWrap");
  const body=document.getElementById("drxBody");
  const drxImg=IMG_DRX[code];
  if(drxImg) imgWrap.innerHTML=`<img src="${drxImg}" alt="Difractograma ${code}">`;
  else imgWrap.innerHTML=`<div class="drx-img-placeholder">&#128300; Difractograma no disponible.<br>Guarda como <b>${code}_drx.png</b> en <b>imagenes_drx/</b></div>`;
  const d=DRX_DATA[code];
  if(!d||(!d.minerales.length&&!d.descripcion)){
    body.innerHTML=`<div class="no-drx">No hay datos DRX EVA para esta muestra.</div>`;return;}
  let tableRows="";
  const maxLen=Math.max(d.cod_ids.length,d.minerales.length,d.formulas.length);
  for(let i=0;i<maxLen;i++){
    tableRows+=`<tr><td>${esc(d.cod_ids[i]||"")}</td><td><b>${esc(d.minerales[i]||"")}</b></td>
      <td style="font-family:monospace;font-size:.75rem">${esc(d.formulas[i]||"")}</td></tr>`;}
  let html="";
  if(tableRows) html+=`<div class="section-title">Fases mineralógicas</div>
    <table class="mineral-table"><thead><tr><th>COD ID</th><th>Mineral</th><th>Fórmula</th></tr></thead><tbody>${tableRows}</tbody></table>`;
  if(d.nota) html+=`<div class="section-title">Nota DRX EVA</div><div class="nota-box">${esc(d.nota)}</div>`;
  if(d.descripcion) html+=`<div class="section-title">Descripción propia</div><div class="description-box">${esc(d.descripcion)}</div>`;
  if(d.interpretacion) html+=`<div class="section-title">Posible interpretación con IA <span class="ia-badge">IA</span></div><div class="interpretacion-box">${esc(d.interpretacion)}</div>`;
  // Geoquímica en tab DRX
  const g=geoByCode[code];
  if(g){
    const rows=SHOW_COLS.filter(c=>g[c]!==undefined&&g[c]!==null).map(c=>
      `<tr><td><b>${c}</b></td><td>${Number(g[c]).toFixed(2)}</td><td>${GEO_UNITS[c]||''}</td></tr>`).join("");
    if(rows) html+=`<div class="section-title">Geoquímica</div>
      <table class="mineral-table"><thead><tr><th>Elemento</th><th>Valor</th><th>Unidad</th></tr></thead>
      <tbody>${rows}</tbody></table>`;
  }
  body.innerHTML=html;
}

function activarTab(tabId){
  document.querySelectorAll(".tab-btn").forEach(b=>b.classList.toggle("active",b.dataset.tab===tabId));
  document.querySelectorAll(".tab-content").forEach(c=>c.classList.toggle("active",c.id===tabId));
}
document.querySelectorAll(".tab-btn").forEach(b=>b.addEventListener("click",()=>activarTab(b.dataset.tab)));
document.getElementById("closeModal").addEventListener("click",()=>
  document.getElementById("modalOverlay").classList.remove("open"));
document.getElementById("modalOverlay").addEventListener("click",(e)=>{
  if(e.target===document.getElementById("modalOverlay"))
    document.getElementById("modalOverlay").classList.remove("open");});
document.addEventListener("keydown",(e)=>{
  if(e.key==="Escape") document.getElementById("modalOverlay").classList.remove("open");});

/* ============================================================
   PAGINA DRX Y GEOQUIMICA
   ============================================================ */
const GEO_UNITS = {
  SiO2:'%',Al2O3:'%','Fe2O3(T)':'%',MnO:'%',MgO:'%',CaO:'%',
  Na2O:'%',K2O:'%',TiO2:'%',P2O5:'%',LOI:'%',Total:'%',
  Ba:'ppm',Sr:'ppm',Y:'ppm',Sc:'ppm',Zr:'ppm',Be:'ppm',V:'ppm'
};
const SHOW_COLS=['SiO2','Al2O3','Fe2O3(T)','MnO','MgO','CaO','Na2O','K2O','TiO2','P2O5','LOI','Total','Ba','Sr','Y','Zr','V','Sc'];

function renderGcPage(){
  renderGcSampleList();
  renderDrxGeoContent(document.getElementById("gc-filter-tipo").value);
}

function renderGcSampleList(){
  const filterTipo = document.getElementById("gc-filter-tipo").value;
  const list = document.getElementById("gc-sample-list");
  let items = samples.filter(s => {
    if(filterTipo && (s.TIPO_ROCA||"").toUpperCase() !== filterTipo) return false;
    return nonEmpty(s.DRX) || geoByCode[s.UNINORTE_CODE];
  });
  list.innerHTML = items.map(s => {
    const col = colorForRock(s.TIPO_ROCA);
    const hasDRX = nonEmpty(s.DRX) ? ' 🔬' : '';
    const hasGeo = geoByCode[s.UNINORTE_CODE] ? ' 📊' : '';
    return `<div class="gc-sample-item" data-code="${s.UNINORTE_CODE}" onclick="scrollToDrxCard('${s.UNINORTE_CODE}')">
      <div class="gc-code" style="color:${col}">⬤ ${s.UNINORTE_CODE}${hasDRX}${hasGeo}</div>
      <div class="gc-tipo">${(s.TIPO_ROCA||"—").toLowerCase()} — ${esc(s.CLASIFICACION||"")}</div>
    </div>`;
  }).join("") || '<div style="padding:16px;color:#888;font-size:.82rem">Sin muestras.</div>';
}

function renderDrxGeoContent(filterTipo){
  const content = document.getElementById("gc-content");
  const items = samples.filter(s => {
    if(filterTipo && (s.TIPO_ROCA||"").toUpperCase() !== filterTipo) return false;
    return nonEmpty(s.DRX) || geoByCode[s.UNINORTE_CODE];
  });
  if(!items.length){ content.innerHTML='<div class="gc-no-data">Sin muestras para este filtro.</div>'; return; }
  content.innerHTML = items.map(s => {
    const code = s.UNINORTE_CODE;
    const g = geoByCode[code];
    const drxImg = IMG_DRX[code];
    const d = DRX_DATA[code];
    const col = colorForRock(s.TIPO_ROCA);

    const imgHtml = drxImg
      ? `<img src="${drxImg}" alt="${code}" style="max-width:100%;border-radius:6px;cursor:zoom-in" onclick="abrirImgFull('${drxImg}','${code}')">`
      : `<div class="drx-geo-img-placeholder">Sin difractograma</div>`;

    let popupBtn = '';
    if(d && (d.minerales.length || d.descripcion || d.interpretacion)){
      popupBtn = `<button class="drx-geo-popup-btn" onclick="abrirMineralPopup('${code}')">
        🔬 Ver fases mineralógicas e interpretación</button>`;
    }

    let geoHtml = '';
    if(g){
      const rows = SHOW_COLS.filter(c=>g[c]!==undefined&&g[c]!==null).map(c=>{
        const val = Number(g[c]);
        const unit = GEO_UNITS[c]||'';
        const isOxide = unit==='%';
        const barW = isOxide ? Math.min(100, val*2) : Math.min(100, val/500*100);
        const barColor = isOxide ? '#1d3557' : '#e63946';
        return `<tr><td><b>${c}</b></td>
          <td><div class="geo-bar-cell">
            <div class="geo-bar" style="width:${barW.toFixed(1)}%;background:${barColor}"></div>
            <span>${val.toFixed(2)}</span>
          </div></td>
          <td style="color:#888;font-size:.72rem">${unit}</td></tr>`;
      }).join("");
      geoHtml = `<div style="margin-top:8px"><b style="font-size:.8rem;color:#1d3557">Geoquímica</b>
        <table class="geo-table-visual"><thead><tr><th>Elemento</th><th>Valor</th><th>Und.</th></tr></thead>
        <tbody>${rows}</tbody></table></div>`;
    }

    return `<div class="drx-geo-card" id="drxcard-${code}">
      <div class="drx-geo-card-header">
        <span class="drx-geo-code" style="color:${col}">⬤ ${code}</span>
        <span class="drx-geo-tipo">${s.TIPO_ROCA||"—"} — ${esc(s.CLASIFICACION||"")}</span>
        <button class="btn-detalle" style="margin-left:auto" onclick="abrirFicha('${code}')">Ver ficha completa →</button>
      </div>
      <div class="drx-geo-body">
        <div class="drx-geo-img">${imgHtml}${popupBtn}</div>
        <div class="drx-geo-info">${geoHtml||'<div class="no-drx" style="padding:16px;color:#888">Sin datos geoquímicos.</div>'}</div>
      </div>
    </div>`;
  }).join("");
}

function scrollToDrxCard(code){
  document.querySelectorAll(".gc-sample-item").forEach(el=>
    el.classList.toggle("selected", el.dataset.code===code));
  const card = document.getElementById("drxcard-"+code);
  if(card) card.scrollIntoView({behavior:"smooth", block:"start"});
}

function abrirMineralPopup(code){
  const d = DRX_DATA[code];
  if(!d) return;
  let html=`<button class="mineral-popup-close" onclick="cerrarMineralPopup()">✕</button>
    <h3>🔬 ${code} — DRX EVA</h3>`;
  const maxLen=Math.max((d.cod_ids||[]).length,(d.minerales||[]).length,(d.formulas||[]).length);
  if(maxLen>0){
    let rows='';
    for(let i=0;i<maxLen;i++) rows+=`<tr><td>${esc((d.cod_ids||[])[i]||"")}</td>
      <td><b>${esc((d.minerales||[])[i]||"")}</b></td>
      <td style="font-family:monospace;font-size:.75rem">${esc((d.formulas||[])[i]||"")}</td></tr>`;
    html+=`<div class="section-title">Fases mineralógicas</div>
      <table class="mineral-table"><thead><tr><th>COD ID</th><th>Mineral</th><th>Fórmula</th></tr></thead>
      <tbody>${rows}</tbody></table>`;
  }
  if(d.nota) html+=`<div class="section-title">Nota</div><div class="nota-box">${esc(d.nota)}</div>`;
  if(d.descripcion) html+=`<div class="section-title">Descripción propia</div><div class="description-box">${esc(d.descripcion)}</div>`;
  if(d.interpretacion) html+=`<div class="section-title">Posible interpretación con IA</div><div class="interpretacion-box">${esc(d.interpretacion)}</div>`;
  const overlay=document.createElement("div");
  overlay.className="mineral-popup-overlay";
  overlay.id="mineralPopupOverlay";
  overlay.innerHTML=`<div class="mineral-popup">${html}</div>`;
  overlay.addEventListener("click",e=>{if(e.target===overlay) cerrarMineralPopup();});
  document.body.appendChild(overlay);
}
function cerrarMineralPopup(){
  const el=document.getElementById("mineralPopupOverlay");
  if(el) el.remove();
}
function abrirImgFull(src, title){
  const overlay=document.createElement("div");
  overlay.className="mineral-popup-overlay";
  overlay.innerHTML=`<div style="max-width:90vw;max-height:90vh;position:relative">
    <button class="mineral-popup-close" style="top:-30px;right:0;color:#fff;font-size:1.5rem"
      onclick="this.parentElement.parentElement.remove()">✕</button>
    <img src="${src}" alt="${title}" style="max-width:90vw;max-height:85vh;border-radius:8px;">
  </div>`;
  overlay.addEventListener("click",e=>{if(e.target===overlay) overlay.remove();});
  document.body.appendChild(overlay);
}

document.getElementById("gc-filter-tipo").addEventListener("change", ()=>{
  renderGcSampleList();
  renderDrxGeoContent(document.getElementById("gc-filter-tipo").value);
});

/* ============================================================
   PAGINA LAMINAS DELGADAS
   ============================================================ */
function renderLaminasPage(){
  const content = document.getElementById("lam-page-content");
  const conLaminas = samples.filter(s => (IMG_LAMINAS[s.UNINORTE_CODE]||[]).length > 0);
  if(!conLaminas.length){
    content.innerHTML='<div style="padding:40px;text-align:center;color:#888">No hay láminas delgadas disponibles.</div>';
    return;
  }
  content.innerHTML = conLaminas.map(s => {
    const code = s.UNINORTE_CODE;
    const pairs = IMG_LAMINAS[code] || [];
    const col = colorForRock(s.TIPO_ROCA);

    const pairsHtml = pairs.map((pair, idx) => `
      <div style="margin-bottom:20px">
        <div class="lam-pair-title">Par ${idx+1}</div>
        <div class="lam-pair-grid">
          <div class="lam-pair-item">
            ${pair.nc ? `<img src="${pair.nc}" alt="NC ${idx+1}" onclick="abrirImgFull('${pair.nc}','${code} NC ${idx+1}')">` : '<div class="drx-geo-img-placeholder">Sin imagen NC</div>'}
            <div class="lam-pair-label">NC — Nícoles Cruzados</div>
            <div class="lam-pair-coords">Coordenadas: —</div>
          </div>
          <div class="lam-pair-item">
            ${pair.np ? `<img src="${pair.np}" alt="NP ${idx+1}" onclick="abrirImgFull('${pair.np}','${code} NP ${idx+1}')">` : '<div class="drx-geo-img-placeholder">Sin imagen NP</div>'}
            <div class="lam-pair-label">NP — Nícoles Paralelos</div>
            <div class="lam-pair-coords">Coordenadas: —</div>
          </div>
        </div>
      </div>`).join("");

    return `<div class="lam-page-sample">
      <div class="lam-page-header">
        <span class="lam-page-code" style="color:${col}">⬤ ${code}</span>
        <span class="lam-page-tipo">${s.TIPO_ROCA||"—"} — ${esc(s.CLASIFICACION||"")}</span>
        <button class="btn-detalle" style="margin-left:auto" onclick="abrirFicha('${code}')">Ver ficha →</button>
      </div>
      <div class="lam-page-body">
        <div class="lam-desc-grid">
          <div class="lam-desc-box">
            <label>Descripción general</label>
            <p>${esc(s.OBS_PETRO||"—")}</p>
          </div>
          <div class="lam-desc-box">
            <label>Texturas específicas</label>
            <p>—</p>
          </div>
          <div class="lam-desc-box">
            <label>Minerales</label>
            <p>—</p>
          </div>
        </div>
        ${pairsHtml}
      </div>
    </div>`;
  }).join("");
}

/* ============================================================
   INIT
   ============================================================ */
init();
</script>
</body>
</html>
"""

# ---------------------------------------------------------------------------
# INTEGRACION CON R / GCDkit
# ---------------------------------------------------------------------------
R_EXE = r"C:\Program Files\R\R-4.1.3\bin\Rscript.exe"
R_SCRIPT = SCRIPT_DIR / "generar_diagramas_gcdkit.R"

def exportar_csv_geochem(geochem_data, tipo_roca_map, tmp_dir: Path):
    """Exporta geochem y tipos a CSV para el script R."""
    tmp_dir.mkdir(parents=True, exist_ok=True)

    geo_csv = tmp_dir / "geochem.csv"
    tipo_csv = tmp_dir / "tipos.csv"

    if not geochem_data:
        return None, None

    # geochem.csv
    all_keys = list(geochem_data[0].keys())
    with open(geo_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=all_keys)
        w.writeheader()
        for row in geochem_data:
            w.writerow({k: ("" if v is None else v) for k, v in row.items()})

    # tipos.csv
    with open(tipo_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["UNINORTE_CODE", "TIPO_ROCA"])
        for code, tipo in tipo_roca_map.items():
            w.writerow([code, tipo])

    return geo_csv, tipo_csv


def generar_diagramas_r(geochem_data, tipo_roca_map, diag_dir: Path):
    """
    Llama al script R con GCDkit para generar los diagramas.
    Devuelve dict de rutas relativas {clave: 'imagenes/diagramas/xxx.png'}.
    """
    if not geochem_data:
        return {}

    from pathlib import Path as _Path
    import tempfile

    if not _Path(R_EXE).exists():
        print(f"  Aviso: R no encontrado en {R_EXE}")
        return {}
    if not R_SCRIPT.exists():
        print(f"  Aviso: script R no encontrado: {R_SCRIPT}")
        return {}

    tmp_dir = SCRIPT_DIR / "_tmp_geochem"
    geo_csv, tipo_csv = exportar_csv_geochem(geochem_data, tipo_roca_map, tmp_dir)
    if not geo_csv:
        return {}

    diag_dir.mkdir(parents=True, exist_ok=True)

    print("  Llamando a R + GCDkit...")
    try:
        result = subprocess.run(
            [R_EXE, "--vanilla", str(R_SCRIPT),
             str(geo_csv), str(tipo_csv), str(diag_dir)],
            capture_output=True, text=True, timeout=180
        )
        # Mostrar output de R (filtrar mensajes de carga)
        for line in result.stdout.splitlines():
            if line.strip() and not line.startswith("Loading"):
                print(f"  R: {line}")
        if result.returncode != 0:
            for line in result.stderr.splitlines()[-5:]:
                print(f"  R error: {line}")
    except subprocess.TimeoutExpired:
        print("  Aviso: R tardó demasiado, omitiendo diagramas.")
        return {}
    except Exception as e:
        print(f"  Error llamando R: {e}")
        return {}

    # Leer JSON de rutas generado por R
    json_path = diag_dir / "diagramas_gcdkit.json"
    if not json_path.exists():
        print("  R no generó el archivo de rutas.")
        return {}

    try:
        raw = json_path.read_text(encoding="utf-8")
        rutas_abs = json.loads(raw)
    except Exception as e:
        print(f"  Error leyendo JSON de R: {e}")
        return {}

    # Convertir a rutas relativas al HTML
    rutas_rel = {}
    for key, abs_path in rutas_abs.items():
        p = Path(abs_path)
        if p.exists():
            rutas_rel[key] = f"imagenes/diagramas/{p.name}"

    print(f"  -> {len(rutas_rel)} diagramas GCDkit generados")
    return rutas_rel


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def parse_args():
    args = sys.argv[1:]
    if not args:
        sys.exit(__doc__)

    ruta_excel     = args[0]
    ruta_salida    = "mapa_muestras.html"
    fotos_dir      = None
    new_photos_dir = None
    laminas_dir    = None

    i = 1
    while i < len(args):
        if args[i] == "--fotos" and i+1 < len(args):
            fotos_dir = args[i+1]; i += 2
        elif args[i] == "--new-photos" and i+1 < len(args):
            new_photos_dir = args[i+1]; i += 2
        elif args[i] == "--laminas" and i+1 < len(args):
            laminas_dir = args[i+1]; i += 2
        elif not args[i].startswith("--"):
            ruta_salida = args[i]; i += 1
        else:
            i += 1

    return ruta_excel, ruta_salida, fotos_dir, new_photos_dir, laminas_dir


if __name__ == "__main__":
    ruta_excel, ruta_salida, fotos_dir, new_photos_dir, laminas_dir = parse_args()

    excel_path = Path(ruta_excel)
    if not excel_path.exists():
        sys.exit(f"No se encontro el archivo: {ruta_excel}")

    salida_path = Path(ruta_salida).resolve()
    img_out_dir = salida_path.parent / "imagenes"
    diag_dir    = img_out_dir / "diagramas"

    print(f"Leyendo Excel: {ruta_excel}")
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)

    ws_main = wb[SHEET_MAIN] if SHEET_MAIN in wb.sheetnames else wb.worksheets[0]
    filas = leer_hoja_principal(ws_main)
    all_codes = [str(r[0]).strip().upper() for r in filas if r[0]]
    print(f"  -> {len(filas)} muestras en Hoja1")

    drx_data = {}
    if SHEET_DRX in wb.sheetnames:
        print("Leyendo hoja DRX_EVA...")
        drx_data = parsear_drx_eva(wb[SHEET_DRX])
        print(f"  -> {len(drx_data)} muestras con datos DRX EVA")
    else:
        print(f"  Aviso: no se encontro la hoja '{SHEET_DRX}'")

    geochem_data = []
    if SHEET_GEO in wb.sheetnames:
        print("Leyendo hoja Geoquim...")
        geochem_data = leer_geoquim(wb[SHEET_GEO])
        print(f"  -> {len(geochem_data)} muestras con geoquimica")
    else:
        print(f"  Aviso: no se encontro la hoja '{SHEET_GEO}'")

    print("Recopilando imagenes de muestra de mano...")
    img_muestra = recopilar_imagenes_muestra_v2(all_codes, fotos_dir, new_photos_dir, img_out_dir)
    total_con_imgs = sum(1 for v in img_muestra.values() if v)
    print(f"  -> {total_con_imgs} muestras con imagenes")
    sin_imgs = [c for c in all_codes if not img_muestra.get(c)]
    if sin_imgs:
        print(f"  SIN imagenes ({len(sin_imgs)}): {', '.join(sin_imgs)}")

    print("Buscando difractogramas (imagenes_drx/)...")
    img_drx = recopilar_imagenes_drx(all_codes, img_out_dir)
    total_drx = sum(1 for v in img_drx.values() if v)
    print(f"  -> {total_drx} difractogramas encontrados")

    print("Buscando laminas delgadas...")
    img_laminas = recopilar_laminas_delgadas(all_codes, laminas_dir, img_out_dir)

    # Complementar laminas desde drive_urls.json si existen
    drive_map_file = Path(__file__).parent / "drive_urls.json"
    if drive_map_file.exists():
        drive_map = json.loads(drive_map_file.read_text(encoding="utf-8"))
        pat_nc = re.compile(r'(NYB\d+)_(\d+)NC', re.IGNORECASE)
        pat_np = re.compile(r'(NYB\d+)_(\d+)NP', re.IGNORECASE)
        nc_drive, np_drive = {}, {}
        for fname, url in drive_map.items():
            m = pat_nc.match(fname)
            if m:
                code = f"NYB{int(m.group(1)[3:]):02d}"
                nc_drive[(code, m.group(2))] = url
            m = pat_np.match(fname)
            if m:
                code = f"NYB{int(m.group(1)[3:]):02d}"
                np_drive[(code, m.group(2))] = url
        keys = set(nc_drive.keys()) | set(np_drive.keys())
        for (code, num) in sorted(keys):
            if code not in img_laminas:
                continue
            already = {p["num"] for p in img_laminas[code]}
            if num not in already:
                img_laminas[code].append({
                    "nc": nc_drive.get((code, num)),
                    "np": np_drive.get((code, num)),
                    "num": num
                })

    # Mapa de tipo de roca por codigo
    tipo_roca_map = {}
    for r in filas:
        code = str(r[0]).strip().upper() if r[0] else ""
        m = re.match(r'NYB(\d+)', code, re.IGNORECASE)
        if m:
            code = f"NYB{int(m.group(1)):02d}"
        tipo_roca_map[code] = str(r[10] or "").strip().upper()

    diagramas = {}
    if geochem_data:
        print("Generando diagramas geoquimicos con R/GCDkit...")
        diagramas = generar_diagramas_r(geochem_data, tipo_roca_map, diag_dir)
        if not diagramas and HAS_MPL:
            print("  Usando matplotlib como alternativa...")
            diagramas = generar_diagramas(filas, geochem_data, tipo_roca_map, diag_dir)
            print(f"  -> {len(diagramas)} diagramas matplotlib generados")

    print(f"\nGenerando HTML -> {ruta_salida}")
    generar_html(filas, drx_data, geochem_data, img_muestra, img_drx,
                 img_laminas, diagramas, ruta_salida)
    print(f"Imagenes en: {img_out_dir}")
    print("Mantene la carpeta 'imagenes/' junto al HTML.")
    if not HAS_PIL:
        print("\nNota: instala Pillow para redimensionar imagenes:")
        print("  pip install Pillow")
