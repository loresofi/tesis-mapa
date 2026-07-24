"""
Exporta los datos geoquimicos al formato nativo de GCDkit (.dat)
Listo para abrir directamente desde GCDkit con File > Load Data.

Uso:
    python exportar_gcdkit.py "ruta\\archivo.xlsx"
"""

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instala con: pip install openpyxl")

if len(sys.argv) < 2:
    sys.exit(__doc__)

EXCEL = Path(sys.argv[1])
if not EXCEL.exists():
    sys.exit(f"No se encontro: {EXCEL}")

wb = openpyxl.load_workbook(EXCEL, data_only=True)

# ---------- Leer Hoja1 (tipo de roca y clasificacion) ----------
ws1 = wb["Hoja1"]
rows1 = list(ws1.iter_rows(values_only=True))
# Columnas: 0=CODE, 10=TIPO_ROCA, 11=CLASIFICACION
tipo_map   = {}
clasif_map = {}
for row in rows1[1:]:
    if not row[0]: continue
    import re
    m = re.match(r'NYB(\d+)', str(row[0]).strip(), re.IGNORECASE)
    if m:
        code = f"NYB{int(m.group(1)):02d}"
        tipo_map[code]   = str(row[10] or "").strip()
        clasif_map[code] = str(row[11] or "").strip()

# ---------- Leer hoja Geoquim ----------
ws2 = wb["Geoquim"]
rows2 = list(ws2.iter_rows(values_only=True))

# Fila 1: nombres de analitos
# Fila 2: unidades
# Datos desde la primera fila con codigo NYB
headers = [str(h).strip() if h else "" for h in rows2[0]]
units_raw = [str(u).strip() if u else "" for u in rows2[1]]

# Oxidos mayores (%) y elementos traza (ppm)
OXIDOS  = {"SiO2","Al2O3","Fe2O3(T)","MnO","MgO","CaO","Na2O","K2O","TiO2","P2O5","LOI","Total"}
TRAZAS  = {"Ba","Sr","Y","Sc","Zr","Be","V"}

# Indices de columnas de datos (excluye col 0 = sample name)
col_indices = []
col_names   = []
col_units   = []
for j, h in enumerate(headers[1:], 1):
    if not h or j >= len(units_raw):
        continue
    u = units_raw[j]
    if h in OXIDOS:
        col_indices.append(j); col_names.append(h); col_units.append("%")
    elif h in TRAZAS:
        col_indices.append(j); col_names.append(h); col_units.append("ppm")

# Leer muestras
import re
muestras = []
for row in rows2[2:]:
    if not row[0]: continue
    m = re.match(r'NYB(\d+)', str(row[0]).strip(), re.IGNORECASE)
    if not m: continue
    code = f"NYB{int(m.group(1)):02d}"
    vals = []
    for j in col_indices:
        v = row[j] if j < len(row) else None
        if v is None or (isinstance(v, str) and '<' in v):
            vals.append("0")
        else:
            try:    vals.append(str(round(float(v), 4)))
            except: vals.append("0")
    muestras.append((code, vals))

if not muestras:
    sys.exit("No se encontraron datos en la hoja Geoquim.")

# ---------- Escribir archivo .dat (formato GCDkit) ----------
# Formato:
#   Fila 1 (cabecera): Samp  Tipo  Clasif  SiO2  TiO2 ...
#   Fila 2 (unidades): end   end   end     %     %    ...
#   Filas de datos:    NYB05 IGNEA Andesita 52.43 ...

out_path = EXCEL.parent / "geoquim_gcdkit.dat"

lineas = []

# Cabecera
lineas.append("\t".join(["Samp", "Tipo", "Clasificacion"] + col_names))

# Unidades
lineas.append("\t".join(["end", "end", "end"] + col_units))

# Datos
for code, vals in muestras:
    tipo   = tipo_map.get(code, "")
    clasif = clasif_map.get(code, "")
    lineas.append("\t".join([code, tipo, clasif] + vals))

out_path.write_text("\n".join(lineas), encoding="utf-8")

print(f"\nArchivo generado: {out_path}")
print(f"Muestras:  {len(muestras)}")
print(f"Columnas:  {', '.join(col_names)}")
print(f"\nEn GCDkit: File > Load Data > selecciona '{out_path.name}'")
print("Las columnas 'Tipo' y 'Clasificacion' aparecen como etiquetas para colorear puntos.")
